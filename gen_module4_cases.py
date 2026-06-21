# -*- coding: utf-8 -*-
"""
SINH 3 FILE EXCEL ĐƠN QUAY ĐẦU THEO 3 KỊCH BẢN (Module 4) — gen_module4_cases.py
================================================================================
Tạo data/module4_S1.xlsx, module4_S2.xlsx, module4_S3.xlsx — mỗi file gồm các
ĐƠN BỔ SUNG CHIỀU VỀ được đặt BÁM SÁT điểm kết thúc tuyến THỰC TẾ của từng kịch
bản (S1/S2/S3) trong dữ liệu demo, kèm "dữ liệu trình tự" (thứ tự ưu tiên ghép).

Mục tiêu: khi chạy "Tối ưu ghép chuyến quay đầu", hệ thống LUÔN tìm được một tỷ
lệ chuyến quay đầu phù hợp THỰC TẾ (~40–50% số tuyến) — phần còn lại dispatcher
nhập/ghép thủ công. Không bao giờ rơi vào trạng thái "chưa tìm được kết quả".

Cách đảm bảo tỷ lệ: với ~45% tuyến có km rỗng cao nhất, sinh 1 đơn chiều về có
điểm lấy NẰM TRONG bán kính cho phép quanh điểm xe rảnh, điểm trả hướng về depot,
đủ tải, đúng loại xe, doanh thu dương -> chắc chắn ghép được.

Chạy:  python3 gen_module4_cases.py
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import hpl_engine as eng
import engine_ext as ext

BASE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE, "data")
STATIC_FILE = os.path.join(DATA_DIR, "1. HPL_AI_Dispatching_Simulated_Data_VRPTW.xlsx")

NAVY = "0B3D91"; SUB = "E8EEF7"; ZEBRA = "F4F7FB"
FONT = "Times New Roman"
thin = Side(style="thin", color="B7C3D6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header dùng đúng bí danh mà hpl_engine.parse_backhaul_workbook nhận diện được.
HEADERS = [
    "Mã đơn bổ sung", "Loại đơn", "Khách hàng", "Điểm lấy", "Điểm trả",
    "Vĩ độ lấy", "Kinh độ lấy", "Vĩ độ trả", "Kinh độ trả",
    "Khung giờ lấy", "Khung giờ trả", "Tải trọng (kg)", "Thể tích (m³)",
    "Loại xe yêu cầu", "Nhiệt độ yêu cầu", "Doanh thu dự kiến (đ)",
    "Thứ tự ưu tiên ghép", "Mức ưu tiên", "Ghi chú dispatcher",
]

SCEN_TITLE = {
    "S1": "KỊCH BẢN 1 — THỪA NĂNG LỰC · ĐƠN QUAY ĐẦU BÁM ĐIỂM XE RẢNH",
    "S2": "KỊCH BẢN 2 — ĐỦ NĂNG LỰC · ĐƠN QUAY ĐẦU 3PL THEO TUYẾN",
    "S3": "KỊCH BẢN 3 — THIẾU NĂNG LỰC · ĐƠN QUAY ĐẦU ƯU TIÊN LỢI NHUẬN",
}
SCEN_RATIO = {"S1": 0.45, "S2": 0.45, "S3": 0.5}   # tỷ lệ tuyến có đơn ghép (40–50%)


def build_orders_for_scenario(d, sid, ratio):
    """Sinh danh sách đơn quay đầu bám điểm kết thúc các tuyến của kịch bản sid.
    Dùng CHUNG `engine_ext.build_return_orders` (một nguồn logic duy nhất)."""
    orders = eng.validate_orders(d["scenarios"][sid])
    routes, _un, _info = ext.solve_hybrid(
        orders, d["fleet"], d.get("depots"), ext.preset_weights("can_bang"), 30, mode="greedy")
    bks = ext.build_return_orders(routes, d["fleet"], ratio=ratio, label=f"Mẫu Module 4 · {sid}")
    rows = []
    for o in bks:
        rows.append([
            f"BK-{sid}-{o['seq']:02d}",
            "Hàng khô", o["customer"], o["pickup_name"], o["delivery_name"],
            o["pickup_lat"], o["pickup_lon"], o["delivery_lat"], o["delivery_lon"],
            "06:00-21:00", "07:00-22:00", o["weight_kg"], o["volume_m3"],
            o["min_vehicle"], "Thường", o["revenue"],
            o["seq"], ("Cao" if o["seq"] <= len(bks) // 2 else "Trung bình"),
            f"Ghép chiều về (km rỗng cao) — đơn #{o['seq']}",
        ])
    return rows, len(routes), len(bks)


GUIDE = lambda sid, n_target, n_routes: [
    (SCEN_TITLE[sid], True),
    ("", False),
    (f"File đơn quay đầu mẫu cho {sid}: {n_target}/{n_routes} tuyến có đơn ghép chiều về", False),
    ("(tương ứng ~%.0f%% số tuyến) — phần còn lại dispatcher nhập/ghép thủ công." % (100.0 * n_target / max(1, n_routes)), False),
    ("", False),
    ("Cách dùng: Module 4 → 'Nhập file 3PL' → chọn file này → 'Tối ưu ghép chuyến'.", False),
    ("Hoặc bấm 'Nạp đơn quay đầu mẫu' trong Module 4 để nạp đúng file theo kịch bản đang xem.", False),
    ("Cột 'Thứ tự ưu tiên ghép' là dữ liệu trình tự để ưu tiên xử lý.", False),
    ("Bắt buộc có Vĩ độ/Kinh độ điểm lấy & trả để hệ thống tính tuyến và điểm phù hợp.", False),
]


def style_sheet(ws, headers, rows, money_cols=(), title=""):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, size=13, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=j, value=h)
        cell.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[2].height = 30
    for i, row in enumerate(rows, start=3):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = Font(name=FONT, size=11, color="1A1A1A")
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            if j in money_cols and isinstance(val, (int, float)):
                cell.number_format = '#,##0" đ"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
    for j, h in enumerate(headers, 1):
        letter = openpyxl.utils.get_column_letter(j)
        maxlen = max([len(str(h))] + [len(str(r[j - 1])) for r in rows]) if rows else len(str(h))
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 11), 34)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(len(headers))}{max(2, len(rows) + 2)}"


def gen_file(d, sid):
    rows, n_routes, n_target = build_orders_for_scenario(d, sid, SCEN_RATIO[sid])
    wb = openpyxl.Workbook()
    g = wb.active
    g.title = "Huong_dan"
    g.sheet_view.showGridLines = False
    for i, (txt, big) in enumerate(GUIDE(sid, n_target, n_routes), 1):
        cell = g.cell(row=i, column=1, value=txt)
        cell.font = Font(name=FONT, size=(13 if big else 11), bold=big, color=(NAVY if big else "1A1A1A"))
    g.column_dimensions["A"].width = 110
    style_sheet(wb.create_sheet(f"DonQuayDau_{sid}"), HEADERS, rows,
                money_cols=(16,), title=SCEN_TITLE[sid])
    out = os.path.join(DATA_DIR, f"module4_{sid}.xlsx")
    wb.save(out)
    return out, n_target, n_routes


def main():
    d = eng.parse_workbook(open(STATIC_FILE, "rb").read())["data"]
    for sid in ("S1", "S2", "S3"):
        out, n_target, n_routes = gen_file(d, sid)
        print(f"Đã tạo {out}: {n_target}/{n_routes} tuyến có đơn ghép (~{100*n_target/max(1,n_routes):.0f}%)")


if __name__ == "__main__":
    main()
