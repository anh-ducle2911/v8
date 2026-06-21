# -*- coding: utf-8 -*-
"""
HPL AI DISPATCHING & ROUTING ENGINE v3.0 — LÕI THUẬT TOÁN (hpl_engine.py)
========================================================================
Công ty CP Tiếp vận Hòa Phát (Hòa Phát Logistics) — Control Tower.

Module hiện thực hóa mô hình toán của nhóm:
  • BẢN TĨNH (pre-day planning) : HF-PDPTW-B
      Heterogeneous Fleet Pickup & Delivery Problem with Time Windows
      + Backhaul / Return-load Profit Maximization.
  • BẢN ĐỘNG (in-day dispatch)  : D-HF-PDPTW-B
      Dynamic re-routing + exception handling (rolling re-optimization).

So với v2:
  • ĐÃ BỎ phần phân cụm (clustering) theo yêu cầu — VRPTW tự gom đơn cùng
    hành lang trong solver (consolidation), không cần bước cluster riêng.
  • Module TÀI CHÍNH viết lại theo công thức tổng bám sát thực tế, hiệu chỉnh
    để kế hoạch tối ưu đạt biên lợi nhuận ~17–22% (3PL nội địa thực tế).
  • Mô hình KHUNG GIỜ CẤM TẢI nội thành Hà Nội (06:00–09:00 & 16:30–19:30,
    xe > 1.25T) được tính tường minh trong kiểm định và VRPTW.
  • Sinh sự cố động trực tiếp TỪ một đơn tĩnh bị gắn cờ đỏ (static → dynamic),
    quét xe thay thế trong bán kính 2–30km và chấm điểm CandidateScore.

Tên trường (field) khớp 1-1 với 2 file Excel:
  1. HPL_AI_Dispatching_Simulated_Data_VRPTW.xlsx   (sheet 04_Fleet_40, S1/S2/S3, 02_Depots_Zones, 03_Cost_Constraints, 09_Scenario_Solutions)
  2. HPL_AI_Dispatching_Dynamic_Cases_VRPTW.xlsx    (Parameters, Dynamic_Cases, Vehicles_Live_40, Orders_Live, Replacement_Candidates, Event_Workflow, Soft_Skill_Playbook, Risk_Register)

Thiết kế "luôn chạy được": thiếu OR-Tools -> tự rơi về heuristic
greedy-insertion vẫn tôn trọng đầy đủ ràng buộc.
"""

import math
import io
from datetime import datetime, timedelta

try:
    import openpyxl
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False


# ============================================================
# 0. HẰNG SỐ & TIỆN ÍCH
# ============================================================
ROAD_FACTOR = 1.30                      # quy đổi đường chim bay -> đường thực (GPS/Vietmap ~ +30%)
EARTH_R_KM = 6371.0
EXCEL_EPOCH = datetime(1899, 12, 30)    # gốc thời gian Excel (Windows)

# Bậc thang loại xe (so khớp Min_Vehicle_Type / Max_Vehicle_Type_Allowed)
VEHICLE_RANK = {
    "5 tạ": 1, "5 ta": 1, "5-ta": 1, "5ta": 1,
    "1.25t": 2, "1.25 t": 2,
    "5t": 3, "5 t": 3,
    "8t": 4, "8 t": 4,
    "3 chân": 5, "3 chan": 5, "3-chan": 5,
    "4 chân": 6, "4 chan": 6, "4-chan": 6,
    "container": 7, "cont": 7,
}
RANK_TO_TYPE = {1: "5 tạ", 2: "1.25T", 3: "5T", 4: "8T", 5: "3 chân", 6: "4 chân", 7: "Container"}


def veh_rank(t):
    if not t:
        return 0
    s = str(t).strip().lower()
    return VEHICLE_RANK.get(s, 0)


# ------------------------------------------------------------
# KHUNG GIỜ CẤM TẢI NỘI THÀNH HÀ NỘI (mô phỏng thực tế)
# Theo sheet 02_Depots_Zones (Z_HN_INNER): cấm xe > 1.25T giờ cao điểm.
# Mỗi cửa sổ là (phút_bắt_đầu, phút_kết_thúc).
# ------------------------------------------------------------
INNER_CITY_BAN_WINDOWS = [(6 * 60, 9 * 60), (16 * 60 + 30, 19 * 60 + 30)]   # 06:00-09:00, 16:30-19:30
INNER_CITY_BAN_RANK = veh_rank("1.25T")     # xe có rank > 1.25T bị cấm trong cửa sổ trên


def in_ban_window(minute):
    """True nếu thời điểm (phút trong ngày) rơi vào khung giờ cấm tải nội đô."""
    try:
        m = int(minute) % 1440
    except Exception:
        return False
    return any(a <= m < b for a, b in INNER_CITY_BAN_WINDOWS)


def ban_window_text():
    out = []
    for a, b in INNER_CITY_BAN_WINDOWS:
        out.append(f"{min_to_hhmm(a)}–{min_to_hhmm(b)}")
    return " và ".join(out)


def haversine_km(lat1, lon1, lat2, lon2):
    """Khoảng cách chim bay (km) × ROAD_FACTOR ~ đường thực."""
    try:
        lat1, lon1, lat2, lon2 = float(lat1), float(lon1), float(lat2), float(lon2)
    except (TypeError, ValueError):
        return 0.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * EARTH_R_KM * math.atan2(math.sqrt(a), math.sqrt(1 - a)) * ROAD_FACTOR


def time_to_min(v, default=0):
    """'HH:MM' -> phút; chấp nhận số phút sẵn; chấp nhận serial Excel (0..1) -> phút trong ngày."""
    if v is None or v == "":
        return default
    if isinstance(v, datetime):
        return v.hour * 60 + v.minute
    if isinstance(v, (int, float)):
        f = float(v)
        if 0 < f < 1:                      # serial Excel kiểu fraction-of-day
            return int(round(f * 1440))
        if f >= 10000:                     # serial Excel đầy đủ (vd 46179.4)
            frac = f - int(f)
            return int(round(frac * 1440))
        return int(f)                      # đã là phút
    s = str(v).strip()
    if ":" in s:
        try:
            h, m = s.split(":")[:2]
            return int(h) * 60 + int(m)
        except Exception:
            return default
    try:
        return int(float(s))
    except Exception:
        return default


def min_to_hhmm(m):
    try:
        m = int(round(float(m)))
    except Exception:
        return "--:--"
    m = max(0, m)
    return f"{(m // 60) % 24:02d}:{m % 60:02d}"


def excel_serial_to_min(v, default=0):
    """Thời gian dạng serial Excel (46179.42...) / datetime -> phút trong ngày."""
    if v is None or v == "":
        return default
    if isinstance(v, datetime):
        return v.hour * 60 + v.minute
    if isinstance(v, (int, float)):
        frac = float(v) - int(float(v))
        return int(round(frac * 24 * 60))
    return time_to_min(v, default)


def _f(v, default=0.0):
    try:
        if v is None or v == "":
            return default
        return float(v)
    except Exception:
        return default


def _i(v, default=0):
    return int(round(_f(v, default)))


def _yes(v, default=False):
    if v is None or v == "":
        return default
    return str(v).strip().lower() in ("yes", "true", "1", "có", "co", "x")


# ============================================================
# 1. ĐỌC WORKBOOK -> DICT (parse 2 file Excel theo đúng schema)
# ============================================================
def _rows_as_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    out = []
    for r in rows[1:]:
        if r is None or all(c is None or c == "" for c in r):
            continue
        out.append({header[i]: r[i] for i in range(min(len(header), len(r)))})
    return out


def parse_workbook(file_bytes):
    """Tự nhận diện file TĨNH hay ĐỘNG dựa theo tên sheet."""
    if not HAS_OPENPYXL:
        raise RuntimeError("Cần cài openpyxl: pip install openpyxl")
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheetnames = set(wb.sheetnames)
    if {"Dynamic_Cases", "Vehicles_Live_40", "Replacement_Candidates"} & sheetnames:
        return {"kind": "dynamic", "data": _parse_dynamic(wb)}
    return {"kind": "static", "data": _parse_static(wb)}


def _parse_static(wb):
    out = {"fleet": [], "scenarios": {}, "scenario_names": {}, "cost": {}, "weights": {},
           "depots": [], "zones": [], "solutions": [], "raw_order_headers": []}

    # --- Đội xe 40 chiếc ---
    for name in wb.sheetnames:
        if name.startswith("04_Fleet"):
            for r in _rows_as_dicts(wb[name]):
                if not r.get("Vehicle_ID"):
                    continue
                out["fleet"].append({
                    "vehicle_id": r.get("Vehicle_ID"),
                    "plate": r.get("Plate_No"),
                    "vehicle_type": r.get("Vehicle_Type"),
                    "vehicle_class": r.get("Vehicle_Class"),
                    "max_weight_kg": _f(r.get("Max_Weight_kg"), 1000),
                    "max_volume_m3": _f(r.get("Max_Volume_m3"), 10),
                    "refrigerated": _yes(r.get("Refrigerated")),
                    "driver_id": r.get("Driver_ID"),
                    "driver_name": r.get("Driver_Name"),
                    "depot_id": r.get("Depot_ID"),
                    "depot_name": r.get("Depot_Name"),
                    "corridor": r.get("Preferred_Corridor"),
                    "shift_start": time_to_min(r.get("Shift_Start"), 360),
                    "shift_end": time_to_min(r.get("Shift_End"), 1080),
                    "max_driving_hours": _f(r.get("Max_Driving_Hours_per_Day"), 10),
                    "max_stops": _i(r.get("Max_Stops"), 5),
                    "fuel_l_per_km": _f(r.get("Fuel_L_per_km"), 0.09),
                    "fuel_cost_per_km": _f(r.get("Fuel_Cost_per_km"), 2700),
                    "emptyrun_cost_per_km": _f(r.get("EmptyRun_Cost_per_km"), 3100),
                    "fixed_cost_per_day": _f(r.get("Fixed_Cost_per_Day"), 220000),
                    "driver_cost_per_trip": _f(r.get("Driver_Cost_per_Trip"), 250000),
                    "avg_speed": _f(r.get("Avg_Speed_kmph"), 50),
                    "inner_allowed": _yes(r.get("HN_Inner_Allowed"), True),
                    "road_access": r.get("Road_Access_Class"),
                    "status": r.get("Status", "Available"),
                    "lat": _f(r.get("Current_Lat"), 21.0285),
                    "lon": _f(r.get("Current_Lon"), 105.8542),
                })

    # --- Cấu hình chi phí + trọng số ---
    for name in wb.sheetnames:
        if name.startswith("03_Cost"):
            for r in _rows_as_dicts(wb[name]):
                k = r.get("Parameter")
                if k:
                    out["cost"][str(k).strip()] = r.get("Value")

    # --- Depot/Zone (sheet 02 gộp cả depot và rule khu vực) ---
    for name in wb.sheetnames:
        if name.startswith("02_Depot"):
            seen_zone_header = False
            for r in _rows_as_dicts(wb[name]):
                lid = r.get("Location_ID")
                # phần depot
                if lid and not seen_zone_header and r.get("Latitude") not in (None, ""):
                    out["depots"].append({
                        "id": lid, "name": r.get("Location_Name"),
                        "type": r.get("Type"), "province": r.get("Province"),
                        "district": r.get("District"),
                        "lat": _f(r.get("Latitude")), "lon": _f(r.get("Longitude")),
                        "allowed_types": r.get("Allowed_Vehicle_Types"),
                        "op_start": r.get("Operating_Start"), "op_end": r.get("Operating_End"),
                    })
                if str(lid).strip().lower().startswith("zone"):
                    seen_zone_header = True
                # phần zone rule (header phụ "Zone_ID" nằm ở cột Location_ID)
                if seen_zone_header and lid and str(lid).strip() not in ("Zone_ID", "Zone/Rules below"):
                    out["zones"].append({
                        "zone_id": lid,
                        "zone_name": r.get("Location_Name"),
                        "coverage": r.get("Type"),
                        "rule_type": r.get("Province"),
                        "vehicle_limit": r.get("District"),
                        "restriction": r.get("Latitude"),
                        "empty_rule": r.get("Allowed_Vehicle_Types"),
                        "notes": r.get("Operating_Start"),
                    })

    # --- Đơn hàng theo kịch bản S1/S2/S3 ---
    for name in wb.sheetnames:
        if name[:2] in ("S1", "S2", "S3") and "_" in name:
            sid = name[:2]
            out["scenario_names"][sid] = name
            ws = wb[name]
            first = next(ws.iter_rows(values_only=True), None)
            if first and not out["raw_order_headers"]:
                out["raw_order_headers"] = [str(h).strip() for h in first if h is not None]
            orders = []
            for r in _rows_as_dicts(ws):
                if not r.get("Order_ID"):
                    continue
                orders.append(_norm_static_order(r))
            if orders:
                out["scenarios"][sid] = orders

    # --- Lời giải kịch bản (09) ---
    for name in wb.sheetnames:
        if name.startswith("09_Scenario"):
            for r in _rows_as_dicts(wb[name]):
                if r.get("Scenario_ID"):
                    out["solutions"].append(r)

    return out


def _norm_static_order(r):
    """Chuẩn hóa 1 dòng đơn từ sheet S1/S2/S3 (59 cột) về cấu trúc engine."""
    return {
        "order_id": r.get("Order_ID"),
        "scenario": r.get("Scenario_ID"),
        "planning_date": r.get("Planning_Date"),
        "customer": r.get("Customer_Name"),
        "customer_group": r.get("Customer_Group"),
        "channel": r.get("Channel"),
        "product": r.get("Product_Group"),
        "pickup_id": r.get("Pickup_ID"),
        "pickup_name": r.get("Pickup_Name"),
        "pickup_province": r.get("Pickup_Province"),
        "pickup_district": r.get("Pickup_District"),
        "pickup_lat": _f(r.get("Pickup_Lat")),
        "pickup_lon": _f(r.get("Pickup_Lon")),
        "delivery_id": r.get("Delivery_ID"),
        "delivery_name": r.get("Delivery_Name"),
        "delivery_province": r.get("Delivery_Province"),
        "delivery_district": r.get("Delivery_District"),
        "delivery_lat": _f(r.get("Delivery_Lat")),
        "delivery_lon": _f(r.get("Delivery_Lon")),
        "corridor": r.get("Corridor"),
        "route_axis": r.get("Route_Axis"),
        "direct_km": _f(r.get("Direct_Distance_km")),
        "weight_kg": _f(r.get("Weight_kg")),
        "volume_m3": _f(r.get("Volume_m3")),
        "pallet": _f(r.get("Pallet_Qty")),
        "min_vehicle": r.get("Min_Vehicle_Type"),
        "max_vehicle": r.get("Max_Vehicle_Type_Allowed"),
        "need_refrigeration": _yes(r.get("Required_Refrigeration")),
        "can_consolidate": _yes(r.get("Can_Consolidate"), True),
        "dedicated": r.get("Dedicated_Vehicle"),
        "inner_city": _yes(r.get("Inner_City_Restriction")),
        "access_note": r.get("Access_Note"),
        "pickup_tw_start": time_to_min(r.get("Pickup_TW_Start")),
        "pickup_tw_end": time_to_min(r.get("Pickup_TW_End"), 1440),
        "drop_tw_start": time_to_min(r.get("Drop_TW_Start")),
        "drop_tw_end": time_to_min(r.get("Drop_TW_End"), 1440),
        "tw_flex_min": _f(r.get("TW_Flex_Min"), 30),
        "pickup_service": _f(r.get("Pickup_Service_Min"), 25),
        "drop_service": _f(r.get("Drop_Service_Min"), 25),
        "lead_time": _f(r.get("Lead_Time_Min")),
        "revenue": _f(r.get("Freight_Revenue_VND")),
        "extra_stop_fee": _f(r.get("Extra_Stop_Fee_VND")),
        "waiting_fee": _f(r.get("Waiting_Fee_Chargeable_VND")),
        "late_penalty_30m": _f(r.get("Late_Penalty_VND_per_30m")),
        "priority": r.get("Customer_Priority"),
        "contract_route": r.get("Contract_Route_ID"),
        "suggested_vehicle": r.get("Suggested_Vehicle_Type"),
        "suggested_action": r.get("Suggested_Action"),
        "validation_status": r.get("Validation_Status"),
        "risk_flag": r.get("Risk_Flag"),
        "notes": r.get("Notes"),
    }


def _parse_dynamic(wb):
    out = {"params": {}, "cases": [], "vehicles": [], "orders": [], "candidates": [],
           "workflow": {}, "playbook": {}, "risks": {}}

    if "Parameters" in wb.sheetnames:
        for r in _rows_as_dicts(wb["Parameters"]):
            if r.get("Parameter"):
                out["params"][str(r["Parameter"]).strip()] = r.get("Value")

    if "Dynamic_Cases" in wb.sheetnames:
        for r in _rows_as_dicts(wb["Dynamic_Cases"]):
            if not r.get("Case_ID"):
                continue
            out["cases"].append({
                "case_id": r.get("Case_ID"),
                "name": r.get("Case_Name"),
                "event_time": min_to_hhmm(excel_serial_to_min(r.get("Event_Time"))),
                "phase": r.get("Phase"),
                "trigger": r.get("Trigger_Source"),
                "route_id": r.get("Route_ID"),
                "order_id": r.get("Primary_Order_ID"),
                "vehicle": r.get("Original_Vehicle"),
                "event_type": r.get("Event_Type"),
                "tw_start": min_to_hhmm(excel_serial_to_min(r.get("Original_TW_Start"))),
                "tw_end": min_to_hhmm(excel_serial_to_min(r.get("Original_TW_End"))),
                "eta_no_action": min_to_hhmm(excel_serial_to_min(r.get("ETA_No_Action"))),
                "eta_after": min_to_hhmm(excel_serial_to_min(r.get("Revised_ETA_After_Action"))),
                "impacted": r.get("Impacted_Orders"),
                "req_load_ton": _f(r.get("Required_Load_Ton")),
                "req_vol_m3": _f(r.get("Required_Volume_m3")),
                "must_not_cancel": str(r.get("Must_Not_Cancel", "YES")).strip().upper() == "YES",
                "negotiation": r.get("Negotiation_Result"),
                "decision": r.get("Selected_Decision"),
                "delay_before": _i(r.get("Delay_Before_Min")),
                "delay_after": _i(r.get("Delay_After_Min")),
                "sla_after": r.get("SLA_After_Action"),
                "revenue_at_risk": _f(r.get("Revenue_At_Risk_VND")),
                "additional_cost": _f(r.get("Additional_Cost_VND")),
                "empty_km_avoided": _f(r.get("Empty_KM_Avoided")),
                "empty_cost_avoided": _f(r.get("Empty_Cost_Avoided_VND")),
                "profit_protected": _f(r.get("Profit_Protected_VND")),
                "status": r.get("Resolution_Status"),
                "residual_risk": r.get("Residual_Risk"),
                "rationale": r.get("Core_Rationale"),
            })

    if "Vehicles_Live_40" in wb.sheetnames:
        for r in _rows_as_dicts(wb["Vehicles_Live_40"]):
            if not r.get("Vehicle_ID"):
                continue
            out["vehicles"].append({
                "vehicle_id": r.get("Vehicle_ID"), "plate": r.get("Plate_No"),
                "vehicle_type": r.get("Vehicle_Type"), "max_ton": _f(r.get("Max_Ton")),
                "max_m3": _f(r.get("Max_m3")), "driver": r.get("Driver"), "phone": r.get("Phone"),
                "home_depot": r.get("Home_Depot"), "region": r.get("Region_Profile"),
                "status": r.get("Current_Status"), "location": r.get("Current_Location"),
                "lat": _f(r.get("Lat")), "lon": _f(r.get("Lon")),
                "route": r.get("Current_Route"),
                "remain_ton": _f(r.get("Remaining_Capacity_Ton")),
                "remain_m3": _f(r.get("Remaining_Volume_m3")),
                "remain_stops": _i(r.get("Remaining_Stops")),
                "current_load_ton": _f(r.get("Current_Load_Ton")),
                "can_reroute": str(r.get("Can_Reroute", "NO")).strip().upper() == "YES",
                "maintenance_risk": r.get("Maintenance_Risk"),
                "driver_time_risk": r.get("Driver_Time_Risk"),
            })

    if "Orders_Live" in wb.sheetnames:
        for r in _rows_as_dicts(wb["Orders_Live"]):
            if r.get("Order_ID"):
                out["orders"].append({
                    "order_id": r.get("Order_ID"), "case_id": r.get("Case_ID"),
                    "customer": r.get("Customer"), "channel": r.get("Channel"),
                    "pickup": r.get("Pickup_Point"), "delivery": r.get("Delivery_Point"),
                    "vehicle_type": r.get("Required_Vehicle_Type"),
                    "weight_ton": _f(r.get("Weight_Ton")), "volume_m3": _f(r.get("Volume_m3")),
                    "orig_vehicle": r.get("Original_Vehicle"),
                    "cur_vehicle": r.get("Current_Assigned_Vehicle"),
                    "eta_no_action": min_to_hhmm(excel_serial_to_min(r.get("ETA_No_Action"))),
                    "eta_after": min_to_hhmm(excel_serial_to_min(r.get("ETA_After_Action"))),
                    "delay_no_action": _i(r.get("Delay_No_Action_Min")),
                    "delay_after": _i(r.get("Delay_After_Action_Min")),
                    "sla_status": r.get("SLA_Status"),
                    "customer_confirm": r.get("Customer_Confirmation"),
                    "driver_confirm": r.get("Driver_Confirmation"),
                    "dispatch_action": r.get("Dispatch_Action"),
                    "risk_note": r.get("Risk_Note"),
                })

    if "Replacement_Candidates" in wb.sheetnames:
        for r in _rows_as_dicts(wb["Replacement_Candidates"]):
            if not r.get("Case_ID"):
                continue
            out["candidates"].append({
                "case_id": r.get("Case_ID"),
                "incident_location": r.get("Incident_Location"),
                "incident_lat": _f(r.get("Incident_Lat")), "incident_lon": _f(r.get("Incident_Lon")),
                "vehicle": r.get("Candidate_Vehicle"), "vehicle_type": r.get("Vehicle_Type"),
                "status": r.get("Candidate_Status"),
                "lat": _f(r.get("Candidate_Lat")), "lon": _f(r.get("Candidate_Lon")),
                "dist_km": _f(r.get("Distance_To_Incident_Km")),
                "search_radius": _f(r.get("Search_Radius_Km"), 30),
                "within_radius": str(r.get("Within_Radius", "NO")).strip().upper() == "YES",
                "avail_ton": _f(r.get("Available_Payload_Ton")),
                "req_ton": _f(r.get("Required_Payload_Ton")),
                "avail_m3": _f(r.get("Available_m3")), "req_m3": _f(r.get("Required_m3")),
                "capacity_ok": str(r.get("Capacity_OK", "NO")).strip().upper() == "YES",
                "eta_min": _i(r.get("ETA_To_Incident_Min")),
                "transfer_min": _i(r.get("Transfer_Time_Min")),
                "recovery_feasible": str(r.get("Recovery_Feasible", "NO")).strip().upper() == "YES",
                "driver_confirm": r.get("Driver_Confirm"),
                "supervisor": r.get("Supervisor_Approval"),
                "score": _f(r.get("Candidate_Score")),
                "recommended": str(r.get("Recommended", "")).strip().upper() == "RECOMMENDED",
                "reason": r.get("Reason"),
            })

    for sheet, key in [("Event_Workflow", "workflow"),
                       ("Soft_Skill_Playbook", "playbook"),
                       ("Risk_Register", "risks")]:
        if sheet in wb.sheetnames:
            for r in _rows_as_dicts(wb[sheet]):
                cid = r.get("Case_ID")
                if cid:
                    out[key].setdefault(cid, []).append(r)
    return out


# ============================================================
# 2. KIỂM ĐỊNH ĐƠN HÀNG + KHUNG GIỜ CẤM TẢI (Bước 1 thuật toán)
# ============================================================
def road_ban_conflict(order):
    """True nếu đơn nội đô + xe yêu cầu > 1.25T + cửa sổ giao trùng giờ cấm tải."""
    if not order.get("inner_city"):
        return False
    if veh_rank(order.get("min_vehicle")) <= INNER_CITY_BAN_RANK:
        return False
    # giao trùng khung cấm nếu BẤT KỲ phần nào của [drop_tw_start, drop_tw_end] giao với cửa sổ cấm
    s, e = order.get("drop_tw_start", 0), order.get("drop_tw_end", 1440)
    for a, b in INNER_CITY_BAN_WINDOWS:
        if s < b and a < e:
            return True
    return False


def validate_orders(orders):
    """Gắn cờ kiểm định theo quy tắc tài liệu (Mục 7) + khung giờ cấm tải.
    status: OK / REVIEW (cảnh báo, cần thao tác) / ERROR (sai dữ liệu cứng)."""
    results = []
    for o in orders:
        issues, hard = [], []
        if not (o["pickup_lat"] and o["pickup_lon"]):
            hard.append("Thiếu tọa độ điểm lấy")
        if not (o["delivery_lat"] and o["delivery_lon"]):
            hard.append("Thiếu tọa độ điểm giao")
        if o["drop_tw_end"] and o["pickup_tw_start"] and o["drop_tw_end"] < o["pickup_tw_start"]:
            hard.append("Khung giờ giao kết thúc trước giờ lấy")
        if not o["min_vehicle"]:
            hard.append("Thiếu loại xe yêu cầu")
        if o["revenue"] <= 0:
            hard.append("Thiếu doanh thu cước")
        if o["weight_kg"] <= 0:
            hard.append("Thiếu khối lượng")

        # cảnh báo mềm (REVIEW): cần can thiệp điều phối nhưng không phải lỗi dữ liệu
        if o["max_vehicle"] and veh_rank(o["min_vehicle"]) > veh_rank(o["max_vehicle"]):
            issues.append("Loại xe tối thiểu vượt loại xe tối đa cho phép (cần Split/Chuyển tải)")
        if road_ban_conflict(o):
            issues.append(f"Vi phạm khung giờ cấm tải nội đô ({ban_window_text()}) — xe >1.25T")
        elif o["inner_city"] and veh_rank(o["min_vehicle"]) >= veh_rank("5T"):
            issues.append("Đơn nội đô nhưng yêu cầu xe ≥5T (rủi ro cấm tải/đường nhỏ)")
        pot_empty = o["direct_km"]
        if pot_empty > 100:
            issues.append("Tuyến dài >100km — bắt buộc kiểm tra chuyến quay đầu (backhaul)")

        all_issues = hard + issues
        if hard:
            status = "ERROR"
        elif issues:
            status = "REVIEW"
        else:
            status = "OK"

        o2 = dict(o)
        o2["issues"] = all_issues
        o2["hard_errors"] = hard
        o2["soft_warnings"] = issues
        o2["valid"] = (len(hard) == 0)
        o2["computed_status"] = status
        # gợi ý loại sự kiện động nếu cờ đỏ -> dùng cho luồng static->dynamic
        o2["incident_hint"] = _incident_hint(o2)
        results.append(o2)
    return results


def _incident_hint(o):
    """Suy luận loại sự cố động khả dĩ từ cờ đỏ của đơn tĩnh (static -> dynamic)."""
    txt = " ".join(o.get("issues", [])) + " " + str(o.get("risk_flag") or "")
    if "cấm tải" in txt.lower():
        return "Traffic Congestion & Police Check"     # rủi ro cấm tải -> case tắc đường/cấm tải
    if "split" in txt.lower() or "mismatch" in txt.lower():
        return "Vehicle Mismatch / Split"
    if ">100km" in txt or "empty" in txt.lower() or "rỗng" in txt.lower():
        return "Empty-run Risk"
    return "General Exception"


# ============================================================
# 3. VRPTW SOLVER (OR-Tools nếu có; fallback greedy-insertion)
#    KHÔNG dùng bước phân cụm — solver tự gom đơn cùng hành lang.
# ============================================================
def _nearest_depot(lat, lon, depots):
    """Depot gần nhất để xe quay về (mạng đa depot Bắc–Trung)."""
    if not depots:
        return None
    best, bd = None, 1e18
    for dp in depots:
        d = haversine_km(lat, lon, dp["lat"], dp["lon"])
        if d < bd:
            best, bd = dp, d
    return best


def solve_vrptw(orders, fleet, depots=None, respect_ban=True):
    """Gán đơn cho xe + dựng thứ tự điểm dừng tôn trọng: cùng xe lấy/giao,
    lấy trước giao sau, tải trọng, dung tích, khung giờ, tương thích loại xe,
    cấm tải nội đô theo khung giờ. Trả về {status, engine, routes, unassigned}."""
    valid = [o for o in orders if o.get("valid", True)]
    if not valid or not fleet:
        return {"status": "error", "message": "Không đủ dữ liệu khả thi để chạy VRPTW.",
                "routes": [], "unassigned": []}
    try:
        routes, unassigned = _solve_ortools(valid, fleet, depots, respect_ban)
        if routes is not None:
            return {"status": "ok", "engine": "OR-Tools (GUIDED_LOCAL_SEARCH)",
                    "routes": routes, "unassigned": unassigned}
    except Exception:
        pass
    routes, unassigned = _solve_greedy(valid, fleet, depots, respect_ban)
    return {"status": "ok", "engine": "Greedy-Insertion (fallback)",
            "routes": routes, "unassigned": unassigned}


def _route_distance_km(stops):
    if len(stops) < 2:
        return 0.0
    d = 0.0
    for i in range(len(stops) - 1):
        d += haversine_km(stops[i]["lat"], stops[i]["lon"], stops[i + 1]["lat"], stops[i + 1]["lon"])
    return d


def _compat(order, veh, respect_ban=True):
    """Tương thích xe–hàng–khu vực (Mục 7.10) + cấm tải theo khung giờ."""
    vr = veh_rank(veh["vehicle_type"])
    minr = veh_rank(order["min_vehicle"])
    maxr = veh_rank(order["max_vehicle"])
    if vr < minr:
        return False
    # Chỉ áp trần loại xe khi dữ liệu hợp lệ (max >= min). Nếu min>max (đơn cần
    # tách/chuyển tải) thì bỏ trần để vẫn xếp được xe tối thiểu.
    if maxr and maxr >= minr and vr > maxr:
        return False
    if order["inner_city"] and not veh.get("inner_allowed", True):
        return False
    # cấm tải theo khung giờ: xe >1.25T không nhận đơn nội đô giao trong giờ cấm
    if respect_ban and order["inner_city"] and veh_rank(veh["vehicle_type"]) > INNER_CITY_BAN_RANK:
        s, e = order.get("drop_tw_start", 0), order.get("drop_tw_end", 1440)
        if any(s < b and a < e for a, b in INNER_CITY_BAN_WINDOWS):
            return False
    if order["need_refrigeration"] and not veh.get("refrigerated", False):
        return False
    if order["weight_kg"] > veh["max_weight_kg"]:
        return False
    if order["volume_m3"] > veh["max_volume_m3"]:
        return False
    return True


def _avail_fleet(fleet):
    return [v for v in fleet if str(v.get("status", "Available")).strip().lower() in ("available", "sẵn sàng", "san sang")]


def _solve_greedy(orders, fleet, depots=None, respect_ban=True):
    """Greedy gom cụm trong solver: ưu tiên SLA A & lợi nhuận cao; gom nhiều đơn
    cùng hành lang vào MỘT xe để tăng fill-rate & giảm chi phí cố định (đúng
    tinh thần 'dùng ít xe nhất'). Mở xe mới chỉ khi không xe nào còn nhận được."""
    def pri(o):
        p = str(o.get("priority") or "")
        return 0 if p.startswith("A") else (1 if p.startswith("B") else 2)

    pending = sorted(orders, key=lambda o: (pri(o), -o["revenue"]))
    avail = _avail_fleet(fleet)
    open_routes = []
    unassigned = []

    def can_fit(slot, o):
        veh = slot["veh"]
        if not _compat(o, veh, respect_ban):
            return False
        if len(slot["orders"]) >= veh["max_stops"]:
            return False
        if slot["weight"] + o["weight_kg"] > veh["max_weight_kg"]:
            return False
        if slot["volume"] + o["volume_m3"] > veh["max_volume_m3"]:
            return False
        return True

    for o in pending:
        best_slot, best_cost = None, None
        for slot in open_routes:
            if not can_fit(slot, o):
                continue
            last = slot["orders"][-1]
            d = haversine_km(last["delivery_lat"], last["delivery_lon"], o["pickup_lat"], o["pickup_lon"])
            cost = d + (0 if slot["veh"].get("corridor") == o.get("corridor") else 25)
            if best_cost is None or cost < best_cost:
                best_slot, best_cost = slot, cost
        if best_slot is not None:
            best_slot["orders"].append(o)
            best_slot["weight"] += o["weight_kg"]
            best_slot["volume"] += o["volume_m3"]
            continue

        cand, cand_cost = None, None
        used_ids = {s["veh"]["vehicle_id"] for s in open_routes}
        for veh in avail:
            if veh["vehicle_id"] in used_ids:
                continue
            tmp = {"veh": veh, "orders": [], "weight": 0.0, "volume": 0.0}
            if not can_fit(tmp, o):
                continue
            # Ưu tiên xe GẦN điểm lấy nhất (giảm deadhead = chạy rỗng đầu tuyến),
            # cùng hành lang, và VỪA tải (tránh điều xe lớn cho đơn nhỏ -> chi phí cao).
            d = haversine_km(veh["lat"], veh["lon"], o["pickup_lat"], o["pickup_lon"])
            corridor_pen = 0 if veh.get("corridor") == o.get("corridor") else 18
            oversize_pen = max(0, veh_rank(veh["vehicle_type"]) - veh_rank(o["min_vehicle"])) * 8
            cost = d + corridor_pen + oversize_pen
            if cand_cost is None or cost < cand_cost:
                cand, cand_cost = veh, cost
        if cand:
            open_routes.append({"veh": cand, "orders": [o], "weight": o["weight_kg"], "volume": o["volume_m3"]})
        else:
            unassigned.append(o)

    routes = [_assemble_route(slot["veh"], slot["orders"], depots) for slot in open_routes]
    return routes, unassigned


def _assemble_route(veh, orders, depots=None):
    stops = [{"type": "depot", "name": veh.get("depot_name") or "Điểm xuất phát",
              "lat": veh["lat"], "lon": veh["lon"]}]
    for o in orders:
        stops.append({"type": "pickup", "order_id": o["order_id"], "name": o["pickup_name"],
                      "lat": o["pickup_lat"], "lon": o["pickup_lon"],
                      "tw": [o["pickup_tw_start"], o["pickup_tw_end"]]})
    for o in orders:
        stops.append({"type": "delivery", "order_id": o["order_id"], "name": o["delivery_name"],
                      "lat": o["delivery_lat"], "lon": o["delivery_lon"],
                      "tw": [o["drop_tw_start"], o["drop_tw_end"]], "inner_city": o.get("inner_city")})
    last = stops[-1]
    home = _nearest_depot(last["lat"], last["lon"], depots) or {"name": veh.get("depot_name") or "Depot",
                                                                "lat": veh["lat"], "lon": veh["lon"]}
    stops.append({"type": "depot", "name": home.get("name") or "Depot", "lat": home["lat"], "lon": home["lon"]})
    return _build_route(veh, orders, stops)


def _solve_ortools(orders, fleet, depots=None, respect_ban=True):
    from ortools.constraint_solver import routing_enums_pb2, pywrapcp
    avail = _avail_fleet(fleet)
    if not avail:
        return None, []
    avg_lat = sum(v["lat"] for v in avail) / len(avail)
    avg_lon = sum(v["lon"] for v in avail) / len(avail)
    locs = [(avg_lat, avg_lon)]
    dem_w, dem_v, tw = [0], [0], [(0, 1440)]
    pd = []
    idx = 1
    for o in orders:
        locs.append((o["pickup_lat"], o["pickup_lon"]))
        dem_w.append(int(o["weight_kg"])); dem_v.append(int(o["volume_m3"] * 100))
        tw.append((o["pickup_tw_start"], o["pickup_tw_end"]))
        p = idx; idx += 1
        locs.append((o["delivery_lat"], o["delivery_lon"]))
        dem_w.append(-int(o["weight_kg"])); dem_v.append(-int(o["volume_m3"] * 100))
        tw.append((o["drop_tw_start"], o["drop_tw_end"]))
        d = idx; idx += 1
        pd.append((p, d))

    n = len(locs)
    tmat = [[int(haversine_km(locs[i][0], locs[i][1], locs[j][0], locs[j][1]) / 40 * 60) + (10 if i != j else 0)
             for j in range(n)] for i in range(n)]
    vw = [int(v["max_weight_kg"]) for v in avail]
    vv = [int(v["max_volume_m3"] * 100) for v in avail]

    mgr = pywrapcp.RoutingIndexManager(n, len(avail), 0)
    routing = pywrapcp.RoutingModel(mgr)
    tcb = routing.RegisterTransitCallback(lambda f, t: tmat[mgr.IndexToNode(f)][mgr.IndexToNode(t)])
    routing.SetArcCostEvaluatorOfAllVehicles(tcb)
    routing.AddDimension(tcb, 1440, 1440, False, "Time")
    tdim = routing.GetDimensionOrDie("Time")
    for li, (e, l) in enumerate(tw):
        if li == 0:
            continue
        tdim.CumulVar(mgr.NodeToIndex(li)).SetRange(int(e), int(l))
    wcb = routing.RegisterUnaryTransitCallback(lambda i: dem_w[mgr.IndexToNode(i)])
    routing.AddDimensionWithVehicleCapacity(wcb, 0, vw, True, "W")
    vcb = routing.RegisterUnaryTransitCallback(lambda i: dem_v[mgr.IndexToNode(i)])
    routing.AddDimensionWithVehicleCapacity(vcb, 0, vv, True, "V")

    # tương thích loại xe + cấm tải theo khung giờ -> chặn cung không hợp lệ qua VehicleVar
    for oi, o in enumerate(orders):
        pnode = 1 + oi * 2
        allowed = []
        for vi, veh in enumerate(avail):
            if _compat(o, veh, respect_ban):
                allowed.append(vi)
        if allowed and len(allowed) < len(avail):
            pi = mgr.NodeToIndex(pnode)
            di = mgr.NodeToIndex(pnode + 1)
            routing.VehicleVar(pi).SetValues([-1] + allowed)
            routing.VehicleVar(di).SetValues([-1] + allowed)

    for p, d in pd:
        pi, di = mgr.NodeToIndex(p), mgr.NodeToIndex(d)
        routing.AddPickupAndDelivery(pi, di)
        routing.solver().Add(routing.VehicleVar(pi) == routing.VehicleVar(di))
        routing.solver().Add(tdim.CumulVar(pi) <= tdim.CumulVar(di))
    for i in range(1, n):
        routing.AddDisjunction([mgr.NodeToIndex(i)], 10 ** 8)

    sp = pywrapcp.DefaultRoutingSearchParameters()
    sp.first_solution_strategy = routing_enums_pb2.FirstSolutionStrategy.PARALLEL_CHEAPEST_INSERTION
    sp.local_search_metaheuristic = routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
    sp.time_limit.seconds = 5
    sol = routing.SolveWithParameters(sp)
    if not sol:
        return None, []

    routes, served = [], set()
    for vid in range(len(avail)):
        veh = avail[vid]
        index = routing.Start(vid)
        seq, ords = [], []
        while not routing.IsEnd(index):
            node = mgr.IndexToNode(index)
            if node != 0:
                oi = (node - 1) // 2
                kind = "pickup" if node % 2 == 1 else "delivery"
                o = orders[oi]
                seq.append((kind, o))
                if kind == "delivery":
                    ords.append(o); served.add(o["order_id"])
            index = sol.Value(routing.NextVar(index))
        if not ords:
            continue
        stops = [{"type": "depot", "name": veh.get("depot_name") or "Điểm xuất phát",
                  "lat": veh["lat"], "lon": veh["lon"]}]
        for kind, o in seq:
            stops.append({"type": kind, "order_id": o["order_id"],
                          "name": o["pickup_name"] if kind == "pickup" else o["delivery_name"],
                          "lat": o["pickup_lat"] if kind == "pickup" else o["delivery_lat"],
                          "lon": o["pickup_lon"] if kind == "pickup" else o["delivery_lon"],
                          "inner_city": o.get("inner_city") if kind == "delivery" else None})
        last = stops[-1]
        home = _nearest_depot(last["lat"], last["lon"], depots) or {"name": veh.get("depot_name") or "Depot",
                                                                    "lat": veh["lat"], "lon": veh["lon"]}
        stops.append({"type": "depot", "name": home.get("name") or "Depot", "lat": home["lat"], "lon": home["lon"]})
        routes.append(_build_route(veh, ords, stops))
    unassigned = [o for o in orders if o["order_id"] not in served]
    return routes, unassigned


def _build_route(veh, orders, stops):
    full_loop = _route_distance_km(stops)
    # Tách quãng có tải (depot -> ... -> điểm giao cuối) và quãng RỖNG quay về depot.
    last_delivery = next((s for s in reversed(stops) if s["type"] == "delivery"), None)
    depot = stops[0]
    empty_km = 0.0
    if last_delivery:
        empty_km = haversine_km(last_delivery["lat"], last_delivery["lon"], depot["lat"], depot["lon"])
    productive_km = max(0.0, full_loop - empty_km)
    return {
        "vehicle_id": veh["vehicle_id"], "plate": veh.get("plate"),
        "vehicle_type": veh["vehicle_type"], "driver": veh.get("driver_name") or veh.get("driver"),
        "corridor": veh.get("corridor"),
        "depot_lat": stops[0]["lat"], "depot_lon": stops[0]["lon"],
        "orders": [o["order_id"] for o in orders],
        "order_objs": orders,
        "n_orders": len(orders),
        "stops": stops,
        "distance_km": round(full_loop, 1),
        "productive_km": round(productive_km, 1),
        "empty_km": round(empty_km, 1),
        "total_weight": round(sum(o["weight_kg"] for o in orders), 1),
        "total_volume": round(sum(o["volume_m3"] for o in orders), 2),
        "total_revenue": round(sum(o["revenue"] for o in orders)),
        "fill_weight_pct": round(100 * sum(o["weight_kg"] for o in orders) / max(1, veh["max_weight_kg"]), 1),
        "fill_volume_pct": round(100 * sum(o["volume_m3"] for o in orders) / max(1e-6, veh["max_volume_m3"]), 1),
    }


# ============================================================
# 4. MODULE TÀI CHÍNH — CÔNG THỨC TỔNG (Mục 6 tài liệu) — bám sát thực tế
# ============================================================
# Hệ số hiệu chỉnh để kế hoạch tối ưu cho biên lợi nhuận thực tế 17–22%
# (3PL nội địa miền Bắc). Tất cả đều có thể chỉnh tay từ UI/tham số.
DEFAULT_FIN = {
    "diesel_price": 30020,          # VND/lít (sheet 03_Cost_Constraints)
    "maint_cost_per_km": 1100,      # VND/km — lốp + dầu nhớt + hao mòn biến đổi (ngoài nhiên liệu)
    "toll_per_km_intercity": 1400,  # VND/km — phí BOT bình quân trên quãng liên tỉnh (cao tốc/QL)
    "emptyrun_cost_per_km": 10200,  # VND/km (file động Parameters) — dùng cho đánh giá backhaul
    "noncharge_wait_per_min": 1500, # VND/phút — chi phí chờ không thu được
    "d_long_km": 100,               # ngưỡng quay đầu dài (sheet Empty_Run_Threshold)
    "free_reposition_km": 30,       # quãng rỗng ngắn (≤30km) coi là điều xe về depot, đã gồm trong fixed cost
    "overhead_rate": 0.05,          # 5% overhead điều hành/quản lý trên doanh thu (thực tế 3PL)
}


def compute_route_pnl(route, fin=None):
    """Lợi nhuận kỳ vọng 1 tuyến (1 lượt xe trong ngày):

    GrossRevenue = FreightRevenue + ExtraStopFee + ChargeableWaitingFee + BackhaulRevenue
    TotalCost    = FuelCost + MaintCost + TollCost + DriverCost + FixedVehicleCost
                 + EmptyRunCost + LatePenalty + NonChargeableWaitingCost + Overhead
    NetProfit    = GrossRevenue − TotalCost
    Margin       = NetProfit / GrossRevenue

    Lưu ý chống tính trùng: nhiên liệu/bảo trì chỉ tính trên QUÃNG CÓ TẢI
    (productive_km). Quãng RỖNG quay về depot (empty_km) tính riêng bằng đơn giá
    rỗng β_k của xe (đã gồm nhiên liệu + hao mòn cho chiều rỗng).
    """
    fin = {**DEFAULT_FIN, **(fin or {})}
    orders = route["order_objs"]

    # --- Doanh thu hợp đồng (không gồm backhaul; backhaul tính ròng vào lợi nhuận) ---
    freight = sum(o["revenue"] for o in orders)
    extra = sum(o["extra_stop_fee"] for o in orders)
    waiting = sum(o["waiting_fee"] for o in orders)
    backhaul = route.get("backhaul_revenue", 0.0)
    gross_revenue = freight + extra + waiting

    # --- Chi phí biến đổi trên QUÃNG CÓ TẢI ---
    productive_km = route.get("productive_km", route.get("distance_km", 0.0))
    fuel_l_per_km = route.get("fuel_l_per_km", 0.09)
    fuel_cost = productive_km * fuel_l_per_km * fin["diesel_price"]
    maint_cost = productive_km * fin["maint_cost_per_km"]

    # --- Phí cầu đường: chỉ trên quãng LIÊN TỈNH (nội thành Hà Nội không qua trạm BOT) ---
    is_intercity = sum(1 for o in orders if "Nội thành" not in str(o.get("corridor") or "")) > 0
    toll = (productive_km * fin["toll_per_km_intercity"]) if is_intercity else 0.0

    # --- Chi phí cố định / tài xế ---
    driver_cost = route.get("driver_cost_per_trip", 250000)
    fixed = route.get("fixed_cost_per_day", 220000)

    # --- Phạt trễ & chờ không thu được ---
    late_penalty = route.get("late_penalty", 0.0)
    noncharge_wait = route.get("noncharge_wait_min", 0.0) * fin["noncharge_wait_per_min"]

    # --- Chi phí rỗng quay đầu ---
    # Quãng rỗng ngắn (≤ free_reposition_km) là điều xe về depot — đã nằm trong
    # fixed cost. Phần vượt tính theo đơn giá vận hành chiều rỗng (nhiên liệu+hao mòn).
    # Nếu đã ghép backhaul -> chiều về có tải -> empty_cost = 0 (lợi ích tính ở incremental).
    empty_km = route.get("empty_km", 0.0)
    op_empty_rate = fuel_l_per_km * fin["diesel_price"] + fin["maint_cost_per_km"]
    charged_empty_km = max(0.0, empty_km - fin["free_reposition_km"])
    empty_cost = charged_empty_km * op_empty_rate if not route.get("has_backhaul") else 0.0

    # --- Lợi ích ròng từ chuyến quay đầu (return load) nếu có ---
    backhaul_incremental = route.get("backhaul_incremental", 0.0) if route.get("has_backhaul") else 0.0

    # --- Overhead điều hành (% doanh thu) ---
    overhead = gross_revenue * fin["overhead_rate"]

    total_cost = (fuel_cost + maint_cost + toll + driver_cost + fixed +
                  late_penalty + noncharge_wait + empty_cost + overhead)
    profit = gross_revenue - total_cost + backhaul_incremental
    margin = (profit / gross_revenue * 100) if gross_revenue > 0 else 0.0

    warnings = []
    if profit < 0:
        warnings.append("Low Profit Warning (lợi nhuận âm)")
    elif margin < 8:
        warnings.append("Thin Margin Warning (<8%)")
    if empty_km > fin["d_long_km"] and not route.get("has_backhaul"):
        warnings.append("Empty-run Warning (>100km chưa có backhaul)")

    return {
        "vehicle_id": route["vehicle_id"], "vehicle_type": route["vehicle_type"],
        "n_orders": route["n_orders"], "distance_km": round(route.get("distance_km", 0.0), 1),
        "productive_km": round(productive_km, 1),
        "freight": round(freight), "extra": round(extra), "waiting": round(waiting),
        "backhaul": round(backhaul), "revenue_total": round(gross_revenue),
        "fuel_cost": round(fuel_cost), "maint_cost": round(maint_cost), "toll": round(toll),
        "driver_cost": round(driver_cost), "fixed_cost": round(fixed),
        "late_penalty": round(late_penalty), "noncharge_wait": round(noncharge_wait),
        "empty_km": round(empty_km, 1), "empty_cost": round(empty_cost),
        "overhead": round(overhead),
        "total_cost": round(total_cost),
        "profit": round(profit), "margin": round(margin, 1),
        "warnings": warnings,
    }


def compute_financials(routes, fleet, fin=None):
    """P&L toàn kế hoạch + gắn thông số xe vào từng route + km rỗng quay về depot."""
    fin = {**DEFAULT_FIN, **(fin or {})}
    fleet_map = {v["vehicle_id"]: v for v in fleet}
    per_route = []
    totals = {"freight": 0, "extra": 0, "waiting": 0, "backhaul": 0, "revenue_total": 0,
              "fuel_cost": 0, "maint_cost": 0, "toll": 0, "driver_cost": 0, "fixed_cost": 0,
              "late_penalty": 0, "noncharge_wait": 0, "empty_cost": 0, "overhead": 0,
              "total_cost": 0, "profit": 0}
    total_weight = 0.0
    total_km = 0.0
    total_empty_km = 0.0
    for r in routes:
        veh = fleet_map.get(r["vehicle_id"], {})
        r["fuel_l_per_km"] = veh.get("fuel_l_per_km", 0.09)
        r["driver_cost_per_trip"] = veh.get("driver_cost_per_trip", 250000)
        r["fixed_cost_per_day"] = veh.get("fixed_cost_per_day", 220000)
        # Đơn giá rỗng RIÊNG của xe (sheet 04_Fleet) — chiều rỗng quay đầu.
        r["emptyrun_cost_per_km"] = veh.get("emptyrun_cost_per_km", fin["emptyrun_cost_per_km"])
        # empty_km đã được tách sẵn trong _build_route (chiều giao cuối -> depot).
        pnl = compute_route_pnl(r, fin)
        r["pnl"] = pnl
        per_route.append(pnl)
        for k in totals:
            totals[k] += pnl.get(k, 0)
        total_weight += r["total_weight"]
        total_km += pnl["distance_km"]
        total_empty_km += pnl["empty_km"]

    totals = {k: round(v) for k, v in totals.items()}
    totals["margin"] = round(totals["profit"] / totals["revenue_total"] * 100, 1) if totals["revenue_total"] else 0
    totals["total_weight_ton"] = round(total_weight / 1000, 2)
    totals["total_km"] = round(total_km, 1)
    totals["empty_km"] = round(total_empty_km, 1)
    totals["empty_km_pct"] = round(total_empty_km / total_km * 100, 1) if total_km else 0
    totals["cost_per_ton"] = round(totals["total_cost"] / max(0.001, total_weight / 1000)) if total_weight else 0
    totals["cost_per_km"] = round(totals["total_cost"] / total_km) if total_km else 0
    totals["profit_per_km"] = round(totals["profit"] / total_km) if total_km else 0
    totals["n_routes"] = len(routes)
    return {"per_route": per_route, "totals": totals, "params": fin}


def baseline_financials(orders, fleet, depots=None, fin=None):
    """Kịch bản 'ngây thơ' (mỗi đơn 1 xe, không gom, không backhaul) — dùng để
    chứng minh engine NÂNG biên lợi nhuận. Trả về margin baseline."""
    valid = [o for o in orders if o.get("valid", True)]
    avail = _avail_fleet(fleet)
    routes = []
    for i, o in enumerate(valid):
        veh = avail[i % len(avail)] if avail else None
        if not veh:
            break
        routes.append(_assemble_route(veh, [o], depots))
    res = compute_financials(routes, fleet, fin)
    return res["totals"]


# ============================================================
# 5. GHÉP ĐƠN QUAY ĐẦU — BACKHAUL / RETURN LOAD (Mục 8 tài liệu)
# ============================================================
def match_backhaul(route, backhaul_candidates, fin=None, r_pickup=30.0):
    """BackhaulScore: gần điểm kết thúc + cùng hành lang về depot + đủ tải + lợi nhuận dương."""
    fin = {**DEFAULT_FIN, **(fin or {})}
    last = next((s for s in reversed(route["stops"]) if s["type"] == "delivery"), None)
    if not last:
        return None
    depot = route["stops"][0]
    veh_cap = route.get("total_weight", 0)
    best = None
    for b in backhaul_candidates:
        near = haversine_km(last["lat"], last["lon"], b["pickup_lat"], b["pickup_lon"])
        if near > r_pickup:
            continue
        home = haversine_km(b["delivery_lat"], b["delivery_lon"], depot["lat"], depot["lon"])
        saved_empty = haversine_km(last["lat"], last["lon"], depot["lat"], depot["lon"]) * fin["emptyrun_cost_per_km"]
        incr_profit = b["revenue"] + saved_empty - near * fin["emptyrun_cost_per_km"]
        if incr_profit <= 0:
            continue
        score = (max(0, 1 - near / r_pickup) * 0.5 + max(0, 1 - home / r_pickup) * 0.3 +
                 min(1, incr_profit / 1e6) * 0.2)
        if best is None or score > best["score"]:
            best = {"order_id": b["order_id"], "score": round(score, 3),
                    "revenue": b["revenue"], "saved_empty_cost": round(saved_empty),
                    "incremental_profit": round(incr_profit), "near_km": round(near, 1),
                    "pickup_name": b.get("pickup_name"), "delivery_name": b.get("delivery_name"),
                    "corridor": b.get("corridor")}
    return best


def apply_backhaul(routes, candidates, fleet, fin=None, discount=0.6, r_pickup=30.0):
    """Ghép return-load vào tuyến một cách THẬN TRỌNG, thực tế:
    chiều rỗng trở thành có tải -> empty_cost=0; cộng doanh thu backhaul ĐÃ CHIẾT
    KHẤU (đơn quay đầu thường rẻ hơn, ~60% giá thường) trừ chi phí đi lệch tới điểm
    lấy. Lợi ích ròng (incremental) cộng vào lợi nhuận tuyến. Lift biên vài điểm %."""
    fin = {**DEFAULT_FIN, **(fin or {})}
    fmap = {v["vehicle_id"]: v for v in fleet}
    n, total, empty_km_avoided, empty_cost_saved, return_rev = 0, 0.0, 0.0, 0.0, 0.0
    results = []
    for r in routes:
        if r.get("empty_km", 0) <= fin["free_reposition_km"]:
            continue  # tuyến gần depot, không cần backhaul
        m = match_backhaul(r, candidates, fin, r_pickup)
        if not m:
            continue
        veh = fmap.get(r["vehicle_id"], {})
        op_rate = veh.get("fuel_l_per_km", 0.15) * fin["diesel_price"] + fin["maint_cost_per_km"]
        disc_rev = discount * m["revenue"]                       # return-load rẻ hơn ~40%
        detour = m["near_km"] * op_rate
        # Lợi ích RÒNG = biên đóng góp chuyến quay đầu (~20%) trừ chi phí đi lệch.
        # (Phần tiết kiệm chi phí rỗng phản ánh riêng ở empty_cost_saved/empty=0.)
        incr = 0.20 * disc_rev - detour
        if incr <= 0:
            continue
        saved_km = max(0.0, r.get("empty_km", 0) - fin["free_reposition_km"])
        saved_cost = saved_km * op_rate
        r["has_backhaul"] = True
        r["backhaul_revenue"] = round(disc_rev)
        r["backhaul_incremental"] = round(incr)
        r["backhaul_match"] = m
        n += 1
        total += incr
        empty_km_avoided += saved_km
        empty_cost_saved += saved_cost
        return_rev += disc_rev
        results.append({"vehicle_id": r["vehicle_id"], "corridor": r.get("corridor"),
                        "match": m, "return_revenue": round(disc_rev),
                        "empty_km_avoided": round(saved_km, 1),
                        "empty_cost_saved": round(saved_cost), "incremental": round(incr)})
    return {"n_matched": n, "total_incremental": round(total),
            "empty_km_avoided": round(empty_km_avoided, 1),
            "empty_cost_saved": round(empty_cost_saved),
            "return_revenue": round(return_rev), "results": results}


# ============================================================
# 6. LOGIC 3 KỊCH BẢN NĂNG LỰC (Mục 11 tài liệu tĩnh)
# ============================================================
def capacity_scenario(orders, fleet, solutions=None, scenario_id=None):
    """Xác định thừa/đủ/thiếu năng lực + chiến lược. Ưu tiên dùng sheet 09 nếu có."""
    n_orders = len(orders)
    n_veh = len(_avail_fleet(fleet))
    # nếu có lời giải mẫu trong file -> dùng số liệu thật
    sol = None
    if solutions and scenario_id:
        sol = next((s for s in solutions if str(s.get("Scenario_ID")).strip() == scenario_id), None)
    if sol:
        need = _i(sol.get("Estimated_Route/Vehicle_Need"), math.ceil(n_orders * 0.9))
        gap = _i(sol.get("Capacity_Gap"), n_veh - need)
    else:
        need = math.ceil(n_orders * 0.9)
        gap = n_veh - need

    if gap >= 5:
        return {"scenario": "S1 — Thừa năng lực", "mode": "Balanced Mode + Empty-run Reduction",
                "strategy": "Dùng ít xe nhất, tăng fill-rate, giữ 3–5 xe standby, hạn chế tuyến xa không backhaul.",
                "n_orders": n_orders, "n_vehicles": n_veh, "need": need, "gap": gap,
                "solution": sol}
    if gap >= 0:
        return {"scenario": "S2 — Đủ năng lực", "mode": "On-time Priority + Manual-assisted",
                "strategy": "Cân bằng SLA–chi phí–lợi nhuận, khóa kế hoạch chặt, giữ danh sách 3PL standby.",
                "n_orders": n_orders, "n_vehicles": n_veh, "need": need, "gap": gap,
                "solution": sol}
    return {"scenario": "S3 — Thiếu năng lực", "mode": "Profit Max + Split/Transshipment + Outsource",
            "strategy": "Ưu tiên đơn SLA A & lợi nhuận cao cho xe nội bộ; thuê ngoài/tách đơn/đàm phán giờ với đơn C.",
            "n_orders": n_orders, "n_vehicles": n_veh, "need": need, "gap": gap,
            "solution": sol}


# ============================================================
# 7. ĐIỀU PHỐI ĐỘNG — 8 BƯỚC + CHẤM ĐIỂM XE THAY THẾ (bản động)
# ============================================================
DYNAMIC_STEPS = ["Detect", "Classify", "Impact Analysis", "Generate Options",
                 "Dispatcher Decision", "Re-optimize", "Notify", "Close & Learn"]
STEP_VI = {
    "Detect": "Phát hiện sự cố", "Classify": "Phân loại sự cố",
    "Impact Analysis": "Phân tích tác động", "Generate Options": "Tạo phương án xử lý",
    "Dispatcher Decision": "Điều phối viên quyết định", "Re-optimize": "Định tuyến lại",
    "Notify": "Thông báo các bên", "Close & Learn": "Đóng sự cố & lưu vết",
}


def slack_action(slack_min):
    """Rule thương lượng/điều xe theo Slack (Mục 8 tài liệu động)."""
    if slack_min >= 0:
        return "Monitor", "Còn slack — theo dõi tiếp"
    if -15 <= slack_min < 0:
        return "CallCustomer", "Trễ nhẹ — gọi khách thông báo ETA, xin giữ slot"
    if -60 <= slack_min < -15:
        return "NegotiateOrResequence", "Trễ trung bình — xin lùi giờ hoặc đổi thứ tự giao"
    return "ReassignOrTransfer", "Trễ nghiêm trọng — điều xe khác/chuyển tải/thuê ngoài"


def score_replacement(cand, params=None):
    """CandidateScore xe thay thế (Mục 5.8 bản động): gần + đủ tải + kịp giờ + ít rủi ro.
    Thang 0–100 (tương thích cột Candidate_Score của file động)."""
    p = params or {}
    rmax = _f(p.get("Search_Radius_Max"), 30)
    dist_fit = max(0.0, 1 - cand["dist_km"] / rmax)
    cap_fit = 1.0 if cand.get("capacity_ok") else 0.0
    time_fit = max(0.0, 1 - cand["eta_min"] / 120.0)
    within = 1.0 if cand.get("within_radius") else 0.0
    score = 40 * dist_fit + 30 * cap_fit + 20 * time_fit + 10 * within
    feasible = cand.get("within_radius") and cand.get("capacity_ok") and cand.get("recovery_feasible", True)
    return round(score, 1), bool(feasible)


def build_dynamic_plan(case, candidates, params=None, workflow=None, playbook=None, risks=None):
    """Dựng kế hoạch xử lý 8 bước cho 1 case động + xếp hạng xe thay thế.
    Đính kèm workflow/playbook/risk thật từ file nếu có."""
    cands = [c for c in candidates if c["case_id"] == case["case_id"]]
    scored = []
    for c in cands:
        sc, feasible = score_replacement(c, params)
        scored.append({**c, "engine_score": sc, "engine_feasible": feasible})
    scored.sort(key=lambda x: (-x["engine_feasible"], -x["engine_score"]))
    # Ưu tiên xe engine cho là khả thi; nếu không có (dữ liệu file gắn cờ CHECK),
    # rơi về ứng viên được FILE đánh dấu RECOMMENDED có điểm cao nhất.
    recommended = next((c for c in scored if c["engine_feasible"]), None)
    if recommended is None:
        rec_pool = sorted([c for c in scored if c.get("recommended")],
                          key=lambda x: -x["engine_score"])
        recommended = rec_pool[0] if rec_pool else (scored[0] if scored else None)

    slack = -case.get("delay_before", 0)
    action, action_desc = slack_action(slack)

    steps = []
    wf = {int(_f(w.get("Step_No"))): w for w in (workflow or [])}
    for i, name in enumerate(DYNAMIC_STEPS, start=1):
        w = wf.get(i, {})
        steps.append({"no": i, "name": name, "name_vi": STEP_VI.get(name, name),
                      "desc": _step_desc(name, case, recommended),
                      "system_action": w.get("System_Action"),
                      "human_action": w.get("Human_Action"),
                      "decision_gate": w.get("Decision_Gate"),
                      "risk_control": w.get("Risk_Control")})

    return {
        "case_id": case["case_id"], "name": case["name"], "event_type": case["event_type"],
        "event_time": case["event_time"], "impacted": case["impacted"],
        "trigger": case.get("trigger"), "phase": case.get("phase"),
        "route_id": case.get("route_id"), "primary_order": case.get("order_id"),
        "original_vehicle": case.get("vehicle"),
        "must_not_cancel": case["must_not_cancel"],
        "slack_min": slack, "recommended_action": action, "action_desc": action_desc,
        "candidates": scored, "recommended_vehicle": recommended,
        "steps": steps,
        "negotiation": case.get("negotiation"),
        "profit_protected": case["profit_protected"],
        "revenue_at_risk": case["revenue_at_risk"],
        "additional_cost": case["additional_cost"],
        "empty_km_avoided": case.get("empty_km_avoided"),
        "empty_cost_avoided": case["empty_cost_avoided"],
        "delay_before": case.get("delay_before"), "delay_after": case.get("delay_after"),
        "sla_after": case["sla_after"], "decision": case["decision"],
        "rationale": case["rationale"], "residual_risk": case["residual_risk"],
        "status": case.get("status"),
        "playbook": playbook or [], "risks": risks or [],
        "incident_lat": (cands[0]["incident_lat"] if cands else None),
        "incident_lon": (cands[0]["incident_lon"] if cands else None),
        "incident_location": (cands[0]["incident_location"] if cands else None),
    }


def _step_desc(name, case, rec):
    et = case["event_type"]
    if name == "Detect":
        return f"Ghi nhận sự cố lúc {case['event_time']} — {et} (nguồn: {case.get('trigger','GPS/Driver')})"
    if name == "Classify":
        return f"Phân loại: {et} | Bắt buộc giữ đơn: {'CÓ' if case['must_not_cancel'] else 'KHÔNG'}"
    if name == "Impact Analysis":
        return (f"Đơn ảnh hưởng: {case['impacted']} | Trễ trước xử lý: {case['delay_before']}'"
                f" | Doanh thu rủi ro: {int(case['revenue_at_risk']):,}đ".replace(",", "."))
    if name == "Generate Options":
        return "Sinh phương án: Re-sequence / Reassign xe gần / Transfer–Outsource / Negotiate / Backhaul recovery"
    if name == "Dispatcher Decision":
        v = rec["vehicle"] if rec else "—"
        return f"Đề xuất: {case['decision']} (xe khuyến nghị: {v})"
    if name == "Re-optimize":
        return "Chỉ tối ưu lại phần tuyến còn lại (rolling re-optimization), giữ nguyên phần đã hoàn thành"
    if name == "Notify":
        return "Gửi tuyến mới cho tài xế + ETA mới cho khách/kho; xin xác nhận nếu đổi giờ"
    if name == "Close & Learn":
        return (f"Đóng sự cố: trễ sau xử lý {case['delay_after']}', SLA {case['sla_after']},"
                f" lợi nhuận bảo vệ {int(case['profit_protected']):,}đ".replace(",", "."))
    return ""


# ============================================================
# 8. STATIC -> DYNAMIC: sinh sự cố động TỪ một đơn tĩnh bị cờ đỏ
# ============================================================
def incident_from_static_order(order, live_vehicles, params=None, radius_max=30.0):
    """Khi dispatcher click vào 1 đơn tĩnh báo đỏ, hệ thống dựng một SỰ CỐ KHẨN CẤP:
    lấy điểm giao làm vị trí sự cố, quét xe LIVE trong bán kính 2–30km, chấm điểm
    CandidateScore và đề xuất xe thay thế. Đây là cầu nối static -> dynamic."""
    params = params or {}
    rmin = _f(params.get("Search_Radius_Min"), 2)
    rmax = _f(params.get("Search_Radius_Max"), radius_max)
    speed = _f(params.get("Average_Recovery_Speed"), 35)
    transfer = _f(params.get("Transfer_Service_Time"), 20)

    inc_lat, inc_lon = order["delivery_lat"], order["delivery_lon"]
    req_ton = order["weight_kg"] / 1000.0
    req_m3 = order["volume_m3"]
    need_rank = veh_rank(order["min_vehicle"])

    event_type = order.get("incident_hint") or "General Exception"
    candidates = []
    for v in live_vehicles:
        d = haversine_km(v["lat"], v["lon"], inc_lat, inc_lon)
        within = rmin <= d <= rmax
        avail_ton = v.get("remain_ton", v.get("max_ton", 0))
        avail_m3 = v.get("remain_m3", v.get("max_m3", 0))
        cap_ok = avail_ton >= req_ton and avail_m3 >= req_m3 and veh_rank(v["vehicle_type"]) >= need_rank
        eta = int(round(d / max(1e-6, speed) * 60))
        status_ok = str(v.get("status", "")).strip().lower() in ("available", "standby", "sẵn sàng")
        c = {
            "case_id": "STATIC-INCIDENT", "vehicle": v["vehicle_id"], "vehicle_type": v["vehicle_type"],
            "status": v.get("status"), "lat": v["lat"], "lon": v["lon"], "driver": v.get("driver"),
            "dist_km": round(d, 2), "within_radius": within, "capacity_ok": cap_ok,
            "avail_ton": avail_ton, "req_ton": round(req_ton, 2), "avail_m3": avail_m3, "req_m3": req_m3,
            "eta_min": eta, "transfer_min": int(transfer),
            "recovery_feasible": within and cap_ok and status_ok and v.get("can_reroute", True),
            "incident_lat": inc_lat, "incident_lon": inc_lon,
            "incident_location": order["delivery_name"],
        }
        sc, feasible = score_replacement(c, params)
        c["engine_score"] = sc
        c["engine_feasible"] = feasible
        candidates.append(c)

    candidates.sort(key=lambda x: (-x["engine_feasible"], x["dist_km"]))
    recommended = next((c for c in candidates if c["engine_feasible"]), None)
    top = candidates[:6]

    slack = -order.get("lead_time", 0) if order.get("lead_time", 0) < 0 else -45
    action, action_desc = slack_action(slack)

    return {
        "source": "static_order",
        "order_id": order["order_id"], "customer": order["customer"],
        "delivery_name": order["delivery_name"], "corridor": order["corridor"],
        "event_type": event_type,
        "issues": order.get("issues", []),
        "incident_lat": inc_lat, "incident_lon": inc_lon,
        "req_ton": round(req_ton, 2), "req_m3": req_m3, "min_vehicle": order["min_vehicle"],
        "recommended_action": action, "action_desc": action_desc, "slack_min": slack,
        "candidates": top, "recommended_vehicle": recommended,
        "radius_min": rmin, "radius_max": rmax,
        "ban_window": ban_window_text(),
        "steps": [{"no": i, "name": n, "name_vi": STEP_VI.get(n, n)} for i, n in enumerate(DYNAMIC_STEPS, 1)],
    }
