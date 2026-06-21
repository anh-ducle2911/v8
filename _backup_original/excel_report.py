# -*- coding: utf-8 -*-
"""
XUẤT BÁO CÁO EXCEL (excel_report.py)
====================================
Xuất báo cáo điều phối ra file .xlsx ĐẸP, CHUYÊN NGHIỆP, font Times New Roman:
  • Tiêu đề công ty (navy/đỏ), banner, ngày xuất.
  • Header in đậm, căn giữa, tô màu thương hiệu; border; freeze panes; bộ lọc.
  • Định dạng số tiền, phần trăm, ngày giờ; auto-fit độ rộng cột.
  • Các sheet: Kế hoạch tuyến, Đơn chưa gán, Kiểm định dữ liệu, Tài chính P&L,
    Danh sách sự cố, Nhật ký điều phối, Tổng hợp.
Không xuất CSV — toàn bộ là .xlsx styled.
"""

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HPL_NAVY = "0B3D91"
HPL_RED = "E2231A"
HEADER_BG = "0B3D91"
HEADER_FG = "FFFFFF"
SUBHEAD_BG = "E8EEF7"
TOTAL_BG = "FCE4E4"
ZEBRA = "F4F7FB"

FONT = "Times New Roman"
thin = Side(style="thin", color="B7C3D6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
VND = '#,##0" đ"'
PCT = '0.0"%"'


def _title(ws, text, sub, ncol):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(row=1, column=1, value="CÔNG TY CỔ PHẦN TIẾP VẬN HÒA PHÁT — TRUNG TÂM ĐIỀU PHỐI")
    c.font = Font(name=FONT, size=15, bold=True, color=HPL_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    c = ws.cell(row=2, column=1, value=text)
    c.font = Font(name=FONT, size=12, bold=True, color=HPL_RED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncol)
    c = ws.cell(row=3, column=1, value=sub)
    c.font = Font(name=FONT, size=10, italic=True, color="555555")
    c.alignment = Alignment(horizontal="center", vertical="center")


def _header_row(ws, row, headers):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(name=FONT, size=11, bold=True, color=HEADER_FG)
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 30


def _autofit(ws, start_row, headers, widths=None):
    for j, h in enumerate(headers, start=1):
        letter = get_column_letter(j)
        if widths and j - 1 < len(widths) and widths[j - 1]:
            ws.column_dimensions[letter].width = widths[j - 1]
            continue
        maxlen = len(str(h))
        for row in ws.iter_rows(min_row=start_row, min_col=j, max_col=j):
            v = row[0].value
            if v is not None:
                maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[letter].width = min(max(maxlen + 3, 10), 44)


def _cell(ws, row, col, value, *, money=False, pct=False, center=False, bold=False,
          zebra=False, color=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=11, bold=bold, color=color or "1A1A1A")
    c.border = BORDER
    align = "center" if center else ("right" if (money or pct) else "left")
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=not (money or pct))
    if money:
        c.number_format = VND
    if pct:
        c.number_format = PCT
    if zebra:
        c.fill = PatternFill("solid", fgColor=ZEBRA)
    return c


def _sheet_table(ws, title, sub, headers, rows, widths=None, money_cols=(), pct_cols=(),
                 center_cols=(), total_row=None):
    ws.sheet_view.showGridLines = False
    _title(ws, title, sub, len(headers))
    hr = 5
    _header_row(ws, hr, headers)
    r = hr + 1
    for i, row in enumerate(rows, start=1):
        z = (i % 2 == 0)
        for j, val in enumerate(row, start=1):
            _cell(ws, r, j, val, money=(j in money_cols), pct=(j in pct_cols),
                  center=(j in center_cols), zebra=z)
        r += 1
    if not rows:
        ws.cell(row=r, column=1, value="(Không có dữ liệu)").font = Font(name=FONT, italic=True, color="888888")
        r += 1
    if total_row:
        for j, val in enumerate(total_row, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.font = Font(name=FONT, size=11, bold=True, color=HPL_NAVY)
            c.fill = PatternFill("solid", fgColor=TOTAL_BG)
            c.border = BORDER
            c.alignment = Alignment(horizontal="right" if (j in money_cols or j in pct_cols) else "center",
                                    vertical="center")
            if j in money_cols:
                c.number_format = VND
            if j in pct_cols:
                c.number_format = PCT
    _autofit(ws, hr + 1, headers, widths)
    ws.freeze_panes = f"A{hr + 1}"
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{max(hr, r - 1)}"


def build_report(payload):
    """payload: dict gồm routes, unassigned, validation, financial, incidents, log, meta."""
    meta = payload.get("meta", {})
    stamp = datetime.now().strftime('%d/%m/%Y %H:%M')
    scn = meta.get("scenario", "—")
    wb = openpyxl.Workbook()

    # ---------- 1. KẾ HOẠCH TUYẾN ----------
    ws = wb.active
    ws.title = "Kế hoạch tuyến"
    fin_map = {p["vehicle_id"]: p for p in (payload.get("financial", {}).get("per_route", []))}
    rows = []
    for i, r in enumerate(payload.get("routes", []), start=1):
        p = fin_map.get(r.get("vehicle_id"), {})
        rows.append([i, r.get("vehicle_id"), r.get("vehicle_type"), r.get("driver"),
                     r.get("corridor"), r.get("n_orders"), r.get("distance_km"),
                     r.get("total_weight"), r.get("fill_weight_pct"),
                     p.get("revenue_total"), p.get("total_cost"), p.get("profit"),
                     p.get("margin"), ", ".join(r.get("orders", []))])
    _sheet_table(ws, "Kế hoạch điều phối tuyến",
                 f"Kịch bản: {scn} · Engine: {meta.get('engine','—')} · Xuất lúc: {stamp}",
                 ["STT", "Mã xe", "Loại xe", "Tài xế", "Hành lang", "Số đơn", "Quãng đường (km)",
                  "Tải (kg)", "Tỷ lệ tải %", "Doanh thu", "Chi phí", "Lợi nhuận", "Biên %", "Danh sách đơn"],
                 rows, widths=[6, 12, 10, 16, 16, 8, 14, 11, 10, 14, 14, 14, 9, 38],
                 money_cols=(10, 11, 12), pct_cols=(9, 13), center_cols=(1, 6, 7))

    # ---------- 2. ĐƠN CHƯA GÁN ----------
    ws2 = wb.create_sheet("Đơn chưa gán")
    rows = []
    for i, o in enumerate(payload.get("unassigned", []), start=1):
        rows.append([i, o.get("order_id"), o.get("pickup_name"), o.get("delivery_name"),
                     o.get("tw"), o.get("weight_kg"), o.get("min_vehicle"),
                     o.get("reason"), o.get("suggestion")])
    _sheet_table(ws2, "Danh sách đơn chưa gán được",
                 f"Kịch bản: {scn} · Xuất lúc: {stamp}",
                 ["STT", "Mã đơn", "Điểm lấy", "Điểm trả", "Khung giờ", "Tải (kg)",
                  "Loại xe yêu cầu", "Lý do chưa gán", "Gợi ý xử lý"],
                 rows, widths=[6, 12, 20, 20, 14, 10, 14, 30, 30], center_cols=(1, 6))

    # ---------- 3. KIỂM ĐỊNH DỮ LIỆU ----------
    ws3 = wb.create_sheet("Kiểm định dữ liệu")
    rows = []
    for i, o in enumerate(payload.get("validation", []), start=1):
        rows.append([i, o.get("order_id"), o.get("customer"), o.get("corridor"),
                     o.get("weight_kg"), o.get("min_vehicle"), o.get("drop_tw"),
                     o.get("status_vi"), "; ".join(o.get("issues", [])) or "Hợp lệ",
                     o.get("note", "")])
    _sheet_table(ws3, "Bảng kiểm định dữ liệu đơn hàng",
                 f"Kịch bản: {scn} · Xuất lúc: {stamp}",
                 ["STT", "Mã đơn", "Khách hàng", "Hành lang", "Tải (kg)", "Loại xe",
                  "Khung giao", "Trạng thái", "Lý do cần xem xét", "Ghi chú điều phối"],
                 rows, widths=[6, 12, 18, 16, 10, 12, 14, 14, 34, 24], center_cols=(1, 5))

    # ---------- 4. TÀI CHÍNH P&L ----------
    ws4 = wb.create_sheet("Tài chính P&L")
    fin = payload.get("financial", {})
    tot = fin.get("totals", {})
    rows = []
    for i, p in enumerate(fin.get("per_route", []), start=1):
        rows.append([i, p.get("vehicle_id"), p.get("vehicle_type"), p.get("revenue_total"),
                     p.get("fuel"), p.get("toll"), p.get("driver_cost"), p.get("vehicle_cost"),
                     p.get("handling"), p.get("empty_cost"), p.get("overhead"),
                     p.get("total_cost"), p.get("profit"), p.get("margin")])
    total_row = ["TỔNG", "", "", tot.get("revenue_total"), tot.get("fuel"), tot.get("toll"),
                 tot.get("driver_cost"), tot.get("vehicle_cost"), tot.get("handling"),
                 tot.get("empty_cost"), tot.get("overhead"), tot.get("total_cost"),
                 tot.get("profit"), tot.get("margin")]
    _sheet_table(ws4, "Báo cáo tài chính — Lợi nhuận kỳ vọng (P&L)",
                 f"Biên kế hoạch: {tot.get('margin','—')}% · Sau xử lý sự cố: {tot.get('margin_after','—')}% · Xuất lúc: {stamp}",
                 ["STT", "Mã xe", "Loại xe", "Doanh thu", "Nhiên liệu", "Cầu đường", "Tài xế",
                  "Chi phí xe", "Bốc xếp", "Chạy rỗng", "Quản lý", "Tổng chi phí", "Lợi nhuận", "Biên %"],
                 rows, widths=[6, 12, 10, 15, 13, 12, 13, 13, 12, 12, 12, 15, 15, 9],
                 money_cols=(4, 5, 6, 7, 8, 9, 10, 11, 12, 13), pct_cols=(14,),
                 center_cols=(1,), total_row=total_row)

    # ---------- 5. DANH SÁCH SỰ CỐ ----------
    ws5 = wb.create_sheet("Danh sách sự cố")
    rows = []
    for i, c in enumerate(payload.get("incidents", []), start=1):
        rows.append([i, c.get("case_id"), c.get("order_id"), c.get("event_type"),
                     c.get("priority", "Cao"), c.get("ts", ""), c.get("vehicle", "—"),
                     c.get("route_id", "—"), c.get("status"), c.get("decision", "—")])
    _sheet_table(ws5, "Danh sách sự cố điều phối",
                 f"Xuất lúc: {stamp}",
                 ["STT", "Mã sự cố", "Mã đơn", "Loại sự cố", "Ưu tiên", "Thời điểm",
                  "Xe liên quan", "Tuyến", "Trạng thái", "Phương án chọn"],
                 rows, widths=[6, 14, 12, 20, 10, 18, 14, 12, 14, 30], center_cols=(1, 5))

    # ---------- 6. NHẬT KÝ ĐIỀU PHỐI ----------
    ws6 = wb.create_sheet("Nhật ký điều phối")
    rows = []
    for i, e in enumerate(payload.get("log", []), start=1):
        rows.append([i, e.get("ts"), e.get("case_id"), e.get("source"), e.get("order_id"),
                     e.get("event_type"), e.get("decision"), e.get("vehicle"),
                     e.get("status"), e.get("user", "Điều phối viên"), e.get("note", "")])
    _sheet_table(ws6, "Nhật ký & lưu vết điều phối",
                 f"Phục vụ kiểm toán nội bộ / Close & Learn · Xuất lúc: {stamp}",
                 ["STT", "Thời điểm", "Mã sự cố", "Nguồn", "Đơn", "Loại sự cố", "Phương án",
                  "Xe", "Trạng thái", "Người xử lý", "Ghi chú"],
                 rows, widths=[6, 18, 14, 18, 12, 20, 30, 12, 14, 16, 26], center_cols=(1,))

    # ---------- 7. TỔNG HỢP ----------
    ws7 = wb.create_sheet("Tổng hợp")
    ws7.sheet_view.showGridLines = False
    _title(ws7, "Báo cáo tổng hợp demo điều phối",
           f"Kịch bản: {scn} · Xuất lúc: {stamp}", 2)
    kpis = [
        ("Số tuyến/xe sử dụng", tot.get("n_routes", 0), "num"),
        ("Tổng quãng đường (km)", tot.get("total_km", 0), "num"),
        ("Tỷ lệ chạy rỗng (%)", tot.get("empty_km_pct", 0), "pct"),
        ("Tổng doanh thu", tot.get("revenue_total", 0), "money"),
        ("Tổng chi phí", tot.get("total_cost", 0), "money"),
        ("Lợi nhuận kế hoạch", tot.get("profit", 0), "money"),
        ("Biên lợi nhuận kế hoạch (%)", tot.get("margin", 0), "pct"),
        ("Chi phí phát sinh do sự cố", tot.get("incident_cost", 0), "money"),
        ("Lợi nhuận sau xử lý sự cố", tot.get("profit_after", 0), "money"),
        ("Biên sau xử lý sự cố (%)", tot.get("margin_after", 0), "pct"),
        ("Số đơn chưa gán", len(payload.get("unassigned", [])), "num"),
        ("Số sự cố ghi nhận", len(payload.get("incidents", [])), "num"),
    ]
    r = 5
    for k, v, fmt in kpis:
        a = ws7.cell(row=r, column=1, value=k)
        a.font = Font(name=FONT, size=11, bold=True)
        a.fill = PatternFill("solid", fgColor=SUBHEAD_BG)
        a.border = BORDER
        b = ws7.cell(row=r, column=2, value=v)
        b.font = Font(name=FONT, size=11)
        b.border = BORDER
        b.alignment = Alignment(horizontal="right", vertical="center")
        if fmt == "money":
            b.number_format = VND
        elif fmt == "pct":
            b.number_format = PCT
        r += 1
    ws7.column_dimensions["A"].width = 34
    ws7.column_dimensions["B"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
