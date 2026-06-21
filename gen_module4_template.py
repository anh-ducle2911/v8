# -*- coding: utf-8 -*-
"""
Sinh file Excel mẫu cho Module 4 — Đơn bổ sung / Ghép chuyến quay đầu.
Tạo data/module4_backhaul_orders_template.xlsx gồm 3 sheet (3 trường hợp) +
sheet hướng dẫn. Dữ liệu mẫu đặt GẦN các điểm kết thúc tuyến có km rỗng cao
trong dữ liệu demo S1 để ghép được ngay.

Chạy:  python3 gen_module4_template.py
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(BASE, "data", "module4_backhaul_orders_template.xlsx")

NAVY = "0B3D91"; RED = "E2231A"; SUB = "E8EEF7"; ZEBRA = "F4F7FB"
FONT = "Times New Roman"
thin = Side(style="thin", color="B7C3D6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---- Dữ liệu mẫu (đặt gần điểm kết thúc tuyến rỗng cao của demo S1) ----
S1 = ["Mã đơn bổ sung", "Loại đơn", "Khách hàng", "Điểm lấy", "Điểm trả",
      "Vĩ độ lấy", "Kinh độ lấy", "Vĩ độ trả", "Kinh độ trả",
      "Khung giờ lấy", "Khung giờ trả", "Tải trọng (kg)", "Thể tích (m³)",
      "Loại xe yêu cầu", "Nhiệt độ yêu cầu", "Doanh thu dự kiến (đ)",
      "Chi phí phát sinh (đ)", "Mức ưu tiên", "Ghi chú dispatcher"]
S1_ROWS = [
    ["BK-HP-01", "Hàng khô", "Cảng Hải Phòng JSC", "KCN Đình Vũ - Hải Phòng", "HPL Depot Long Biên",
     20.8550, 106.7300, 21.0262, 105.9142, "12:00-15:00", "15:30-18:30", 950, 4.5, "1.25T", "Thường",
     2200000, 0, "Cao", "Ghép chiều về cho xe vừa giao tại Hải Phòng (rỗng ~101km)"],
    ["BK-HD-02", "Hàng tiêu dùng", "Tân Á Đại Thành", "KCN Đại An - Hải Dương", "Kho Linfox VSIP Bắc Ninh",
     20.9500, 106.3300, 21.0827, 105.9979, "11:00-14:00", "14:30-17:30", 800, 3.8, "1.25T", "Thường",
     1500000, 0, "Trung bình", "Đơn khu vực Hải Dương về Bắc Ninh"],
    ["BK-HY-03", "Hàng khô", "Hòa Phát Hưng Yên", "Nhà máy Mỹ Hào - Hưng Yên", "HPL Depot Long Biên",
     20.9329, 106.0650, 21.0262, 105.9142, "10:30-13:30", "14:00-17:00", 1000, 4.0, "1.25T", "Thường",
     1700000, 0, "Cao", "Tuyến ngắn, ghép nhanh"],
]

S2 = ["Mã đơn 3PL", "Đối tác 3PL", "Điểm lấy", "Điểm trả",
      "Vĩ độ lấy", "Kinh độ lấy", "Vĩ độ trả", "Kinh độ trả",
      "Khung giờ lấy", "Khung giờ trả", "Tải trọng (kg)", "Loại hàng",
      "Loại xe yêu cầu", "Giá nhận đơn (đ)", "Độ lệch tối đa (km)",
      "Điều kiện đặc biệt", "Trạng thái khả dụng", "Ghi chú"]
S2_ROWS = [
    ["3PL-QV-101", "GHN Logistics", "KCN Quế Võ - Bắc Ninh", "HPL Depot Long Biên",
     21.1600, 106.1100, 21.0262, 105.9142, "13:00-16:00", "16:00-19:00", 700, "Hàng khô", "1.25T",
     1450000, 15, "Không", "Khả dụng", "Gần điểm giao Quế Võ"],
    ["3PL-HD-102", "J&T Express", "Big C Hải Dương", "Kho Mỹ Hào Hưng Yên",
     20.9373, 106.3146, 20.9329, 106.0650, "12:30-15:30", "15:30-18:00", 600, "Tiêu dùng", "1.25T",
     1350000, 20, "Không", "Khả dụng", "Đơn lẻ Hải Dương → Hưng Yên"],
    ["3PL-UB-103", "Viettel Post", "TP Uông Bí - Quảng Ninh", "HPL Depot Long Biên",
     21.0315, 106.7669, 21.0262, 105.9142, "13:00-16:00", "17:00-20:00", 850, "Hàng khô", "1.25T",
     1900000, 25, "Không", "Khả dụng", "Ghép xe đang ở Uông Bí (rỗng ~27km)"],
    ["3PL-HB-104", "Ninja Van", "Hoài Đức - Hà Nội", "KCN VSIP Bắc Ninh",
     21.0050, 105.7283, 21.0827, 105.9979, "12:00-15:00", "15:30-18:00", 900, "Tiêu dùng", "1.25T",
     1600000, 18, "Không", "Khả dụng", "Xe kết thúc gần Hoài Đức (rỗng ~184km)"],
]

S3 = ["Mã gợi ý", "Nguồn dữ liệu", "Điểm lấy đề xuất", "Điểm trả đề xuất",
      "Vĩ độ lấy", "Kinh độ lấy", "Vĩ độ trả", "Kinh độ trả",
      "Khung giờ lấy", "Khung giờ trả", "Xe/tuyến có thể ghép", "Tải trọng (kg)",
      "Loại xe yêu cầu", "Khoảng cách lệch tuyến (km)", "Thời gian chờ (phút)",
      "Doanh thu bổ sung (đ)", "Chi phí bổ sung (đ)", "Lợi nhuận ước tính (đ)",
      "Điểm phù hợp", "Lý do đề xuất", "Trạng thái xác nhận"]
S3_ROWS = [
    ["GY-001", "Sàn vận tải nội bộ", "Thái Dương - Thanh Trì", "HPL Depot Long Biên",
     20.9292, 105.8700, 21.0262, 105.9142, "13:30-16:00", "16:00-18:30", "Xe rỗng ~122km", 800, "1.25T",
     8, 15, 1550000, 520000, 1030000, 82, "Gần điểm xe rảnh, giảm chạy rỗng chiều về", "Chờ xác nhận"],
    ["GY-002", "Dữ liệu động website", "Trương Định - Hoàng Mai", "KCN VSIP Bắc Ninh",
     20.9793, 105.8447, 21.0827, 105.9979, "13:00-15:30", "15:30-18:00", "Xe rỗng ~126km", 750, "1.25T",
     10, 20, 1480000, 560000, 920000, 78, "Tận dụng xe sắp về depot Bắc Ninh", "Chờ xác nhận"],
]

GUIDE = [
    ("HƯỚNG DẪN — ĐƠN BỔ SUNG / GHÉP CHUYẾN QUAY ĐẦU (MODULE 4)", True),
    ("", False),
    ("File gồm 3 trường hợp, ứng với 3 tình huống điều phối thực tế:", False),
    ("• Sheet 1_Nhap_tay  — Kịch bản 1: Xe giao xong gần khu công nghiệp, dispatcher nhập tay vài đơn chiều về.", False),
    ("• Sheet 2_Import_3PL — Kịch bản 2: 3PL gửi danh sách nhiều đơn, hệ thống chọn đơn ghép tối ưu nhất.", False),
    ("• Sheet 3_Goi_y_tu_dong — Kịch bản 3: Hệ thống/AI gợi ý đơn phù hợp từ dữ liệu động.", False),
    ("", False),
    ("Cách dùng: Module 1 → ô 'Đơn bổ sung — Ghép chuyến quay đầu' → tải file này lên;", False),
    ("hoặc Module 4 → nút 'Nhập file 3PL'. Sau đó chạy 'Tối ưu ghép chuyến quay đầu'.", False),
    ("Bắt buộc có Vĩ độ/Kinh độ điểm lấy & điểm trả để hệ thống tính được tuyến và điểm phù hợp.", False),
    ("Khung giờ ghi dạng 'HH:MM-HH:MM' (ví dụ 13:00-16:00).", False),
]


def style_sheet(ws, headers, rows, money_cols=(), title=""):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, size=13, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=j, value=h)
        cell.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[2].height = 30
    for i, row in enumerate(rows, start=3):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = Font(name=FONT, size=11, color="1A1A1A")
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            if j in money_cols and isinstance(val, (int, float)):
                cell.number_format = '#,##0" đ"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
    for j, h in enumerate(headers, 1):
        letter = openpyxl.utils.get_column_letter(j)
        maxlen = max([len(str(h))] + [len(str(r[j-1])) for r in rows]) if rows else len(str(h))
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 11), 34)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(len(headers))}{max(2, len(rows)+2)}"


def main():
    wb = openpyxl.Workbook()
    g = wb.active; g.title = "Huong_dan"
    g.sheet_view.showGridLines = False
    for i, (txt, big) in enumerate(GUIDE, 1):
        cell = g.cell(row=i, column=1, value=txt)
        cell.font = Font(name=FONT, size=(13 if big else 11), bold=big, color=(NAVY if big else "1A1A1A"))
    g.column_dimensions["A"].width = 110

    style_sheet(wb.create_sheet("1_Nhap_tay"), S1, S1_ROWS,
                money_cols=(16, 17), title="TRƯỜNG HỢP 1 — ĐƠN QUAY ĐẦU NHẬP TAY (dispatcher)")
    style_sheet(wb.create_sheet("2_Import_3PL"), S2, S2_ROWS,
                money_cols=(14,), title="TRƯỜNG HỢP 2 — ĐƠN QUAY ĐẦU IMPORT TỪ 3PL")
    style_sheet(wb.create_sheet("3_Goi_y_tu_dong"), S3, S3_ROWS,
                money_cols=(16, 17, 18), title="TRƯỜNG HỢP 3 — ĐƠN QUAY ĐẦU GỢI Ý TỰ ĐỘNG")

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("Đã tạo:", OUT)


if __name__ == "__main__":
    main()
