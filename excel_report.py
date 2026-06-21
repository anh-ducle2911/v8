# -*- coding: utf-8 -*-
"""
XUẤT BÁO CÁO & DASHBOARD EXCEL (excel_report.py)
================================================
Tạo workbook .xlsx chuẩn "production / BI-pro":
  • Sheet DASHBOARD: KPI cards, biểu đồ native (auto-update theo dữ liệu nguồn),
    bảng phân tích theo hành lang (pivot bằng công thức), conditional formatting.
  • Các sheet dữ liệu là Excel Table (ListObject) -> có nút lọc, banded rows và
    SẴN SÀNG bấm 1 phát Insert ▸ PivotTable / Slicer / Timeline (openpyxl không tự
    tạo được Pivot/Slicer native — đây là cách tiếp cận đúng & minh bạch).
  • KPI & biểu đồ headline dùng CÔNG THỨC tham chiếu vùng dữ liệu -> tự cập nhật
    khi chỉnh số liệu nguồn.
  • Định dạng tiền VND, phần trăm, freeze panes, màu thương hiệu, spacing gọn.

Không hard-code thời điểm xuất: dùng "Cập nhật lần cuối" động.
"""

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, CellIsRule

# ---- Bảng màu thương hiệu (BI-pro) ----
NAVY = "0B3D91"; NAVY_D = "072A66"; BLUE = "1565C0"; BLUE_L = "EAF1FB"
RED = "E2231A"; RED_D = "B71C1C"; GREEN = "1B7A2F"; GREEN_L = "E8F5EC"
AMBER = "B07A00"; AMBER_L = "FFF6E2"; TEAL = "00838F"; PURPLE = "5C6BC0"
SLATE = "5B6678"; CANVAS = "EEF2F8"; CARD = "FFFFFF"; LINE = "D8E0EC"
INK = "16203A"; SUBHEAD = "E8EEF7"

FONT = "Calibri"
VND = '#,##0" đ"'; PCT = '0.0"%"'; NUM = '#,##0'; NUM1 = '#,##0.0'
thin = Side(style="thin", color=LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _f(v, d=0.0):
    try:
        return float(v)
    except Exception:
        return d


# ============================================================
# TIỆN ÍCH STYLE
# ============================================================
def _fill(hexc):
    return PatternFill("solid", fgColor=hexc)

def _sheet(wb, tab, title, subtitle, headers, rows, *, widths=None, money_cols=(),
           pct_cols=(), num_cols=(), center_cols=(), table_name=None, tab_color=NAVY):
    """Tạo 1 sheet dữ liệu dạng Excel Table. Trả về (ws, header_row, first_data, last_data, col_letters)."""
    ws = wb.create_sheet(tab)
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False
    n = len(headers)
    # tiêu đề
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1, value="CÔNG TY CỔ PHẦN TIẾP VẬN HÒA PHÁT — " + title)
    c.font = Font(name=FONT, size=14, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = Font(name=FONT, size=10, italic=True, color=SLATE)
    hr = 3
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=j, value=h)
        cell.font = Font(name=FONT, size=10.5, bold=True, color="FFFFFF")
        cell.fill = _fill(NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[hr].height = 30
    r = hr + 1
    for row in rows:
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = Font(name=FONT, size=10.5, color=INK)
            cell.border = BORDER
            if j in money_cols:
                cell.number_format = VND
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif j in pct_cols:
                cell.number_format = PCT
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif j in num_cols:
                cell.number_format = NUM
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif j in center_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        r += 1
    last = r - 1
    if not rows:
        ws.cell(row=hr + 1, column=1, value="(Chưa có dữ liệu)").font = Font(name=FONT, italic=True, color="888888")
    # widths
    col_letters = {h: get_column_letter(j) for j, h in enumerate(headers, 1)}
    for j, h in enumerate(headers, 1):
        letter = get_column_letter(j)
        if widths and j - 1 < len(widths) and widths[j - 1]:
            ws.column_dimensions[letter].width = widths[j - 1]
        else:
            maxlen = max([len(str(h))] + [len(str(rows[i][j-1])) for i in range(len(rows))]) if rows else len(str(h))
            ws.column_dimensions[letter].width = min(max(maxlen + 2, 10), 40)
    # Excel Table (chỉ khi có ≥1 dòng dữ liệu)
    if table_name and rows:
        ref = f"A{hr}:{get_column_letter(n)}{last}"
        t = Table(displayName=table_name, ref=ref)
        t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True,
                                          showColumnStripes=False, showFirstColumn=False,
                                          showLastColumn=False)
        ws.add_table(t)
    else:
        ws.auto_filter.ref = f"A{hr}:{get_column_letter(n)}{max(hr, last)}"
    ws.freeze_panes = f"A{hr + 1}"
    return ws, hr, hr + 1, last, col_letters


# ============================================================
# BUILD REPORT
# ============================================================
def build_report(payload):
    meta = payload.get("meta", {})
    kpi = payload.get("kpi", {})
    fin = payload.get("financial", {}) or {}
    stamp = meta.get("generated_display") or datetime.now().strftime("%d/%m/%Y %H:%M")
    scn = meta.get("scenario_vi") or meta.get("scenario", "—")
    fuelc = meta.get("fuel", {}) or {}

    wb = openpyxl.Workbook()
    # Đã BỎ sheet "Dashboard" (KPI cards + biểu đồ) theo yêu cầu — tránh lỗi định
    # dạng khi mở trên các phiên bản Excel/LibreOffice. Chỉ xuất các sheet DỮ LIỆU
    # dạng Excel Table (sẵn sàng tự dựng PivotTable/Slicer) + sheet Hướng dẫn.
    _default_ws = wb.active   # sheet rỗng mặc định -> sẽ xóa ở cuối

    # ---------- SHEET: KẾ HOẠCH TUYẾN ----------
    route_rows = []
    for i, r in enumerate(payload.get("routes", []), 1):
        pnl = r.get("pnl", {}) or {}
        route_rows.append([i, r.get("vehicle_id"), r.get("vehicle_type"), r.get("driver"),
                           r.get("corridor"), r.get("n_orders"), r.get("distance_km"),
                           r.get("empty_km"), r.get("total_weight"), r.get("fill_weight_pct"),
                           pnl.get("revenue_total"), pnl.get("total_cost"), pnl.get("profit"),
                           pnl.get("margin"), "Có" if r.get("has_backhaul") else "—",
                           ", ".join(str(x) for x in (r.get("orders") or []))])
    rh = ["STT", "Mã xe", "Loại xe", "Tài xế", "Tuyến", "Số đơn", "Km", "Km rỗng",
          "Tải (kg)", "Tỷ lệ tải %", "Doanh thu", "Chi phí", "Lợi nhuận", "Biên %",
          "Backhaul", "Danh sách đơn"]
    ws_r, r_hr, r_d0, r_d1, r_cl = _sheet(
        wb, "Kế hoạch tuyến", "KẾ HOẠCH TUYẾN", f"Kịch bản: {scn} · Cập nhật lần cuối: {stamp}",
        rh, route_rows, widths=[5, 11, 9, 15, 15, 7, 8, 9, 10, 10, 14, 14, 14, 8, 9, 34],
        money_cols=(11, 12, 13), pct_cols=(10, 14), num_cols=(7, 8, 9), center_cols=(1, 6, 15),
        table_name="tbl_routes")
    if route_rows:
        ws_r.conditional_formatting.add(f"J{r_d0}:J{r_d1}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=110, color=BLUE))
        ws_r.conditional_formatting.add(f"H{r_d0}:H{r_d1}",
            ColorScaleRule(start_type="min", start_color=GREEN_L, mid_type="percentile", mid_value=50,
                           mid_color="FFE08A", end_type="max", end_color="F4B9B5"))
        ws_r.conditional_formatting.add(f"M{r_d0}:M{r_d1}",
            DataBarRule(start_type="min", end_type="max", color=GREEN))

    # ---------- SHEET: ĐƠN HÀNG ----------
    order_rows = []
    for i, o in enumerate(payload.get("orders", []), 1):
        order_rows.append([i, o.get("order_id"), o.get("customer"), o.get("corridor"),
                           o.get("weight_kg"), o.get("min_vehicle"), o.get("drop_tw"),
                           o.get("revenue"), o.get("status_vi"), o.get("status_gan"),
                           o.get("assigned_vehicle") or "—", "; ".join(o.get("issues", [])) or "Hợp lệ"])
    oh = ["STT", "Mã đơn", "Khách hàng", "Tuyến", "Tải (kg)", "Loại xe", "Khung giao",
          "Doanh thu", "Kiểm định", "Trạng thái gán", "Xe gán", "Lý do cần xem xét"]
    ws_o, o_hr, o_d0, o_d1, o_cl = _sheet(
        wb, "Đơn hàng", "DANH SÁCH ĐƠN HÀNG", f"Kịch bản: {scn} · Cập nhật lần cuối: {stamp}",
        oh, order_rows, widths=[5, 13, 20, 15, 9, 10, 13, 14, 13, 14, 10, 34],
        money_cols=(8,), num_cols=(5,), center_cols=(1, 6), table_name="tbl_orders")
    if order_rows:
        ws_o.conditional_formatting.add(f"J{o_d0}:J{o_d1}",
            CellIsRule(operator="equal", formula=['"Chưa gán"'], fill=_fill("FCE4E4"), font=Font(color=RED_D, bold=True)))
        ws_o.conditional_formatting.add(f"J{o_d0}:J{o_d1}",
            CellIsRule(operator="equal", formula=['"Đã gán"'], fill=_fill(GREEN_L), font=Font(color=GREEN, bold=True)))
        ws_o.conditional_formatting.add(f"I{o_d0}:I{o_d1}",
            CellIsRule(operator="equal", formula=['"Cần xem xét"'], fill=_fill(AMBER_L), font=Font(color=AMBER, bold=True)))

    # ---------- SHEET: XE & TÀI XẾ ----------
    fleet_rows = []
    for i, v in enumerate(payload.get("fleet", []), 1):
        fleet_rows.append([i, v.get("vehicle_id"), v.get("plate"), v.get("vehicle_type"),
                           v.get("driver_name"), v.get("depot_name"), v.get("corridor"),
                           v.get("max_weight_kg"), v.get("max_volume_m3"),
                           "Đang chạy" if v.get("used") else "Standby/Sẵn sàng"])
    fh = ["STT", "Mã xe", "Biển số", "Loại xe", "Tài xế", "Depot", "Tuyến",
          "Tải tối đa (kg)", "Dung tích (m³)", "Trạng thái"]
    ws_f, f_hr, f_d0, f_d1, f_cl = _sheet(
        wb, "Xe & tài xế", "ĐỘI XE & TÀI XẾ", f"Tổng {len(fleet_rows)} xe · Cập nhật lần cuối: {stamp}",
        fh, fleet_rows, widths=[5, 11, 12, 10, 16, 22, 15, 13, 12, 16],
        num_cols=(8, 9), center_cols=(1,), table_name="tbl_fleet")
    if fleet_rows:
        ws_f.conditional_formatting.add(f"J{f_d0}:J{f_d1}",
            CellIsRule(operator="equal", formula=['"Standby/Sẵn sàng"'], fill=_fill(BLUE_L), font=Font(color=NAVY, bold=True)))

    # ---------- SHEET: ĐƠN BỔ SUNG / BACKHAUL ----------
    bk_rows = []
    for i, m in enumerate(payload.get("backhaul_matches", []), 1):
        bk_rows.append([i, m.get("vehicle_id"), m.get("corridor"), (m.get("match") or {}).get("order_id"),
                        (m.get("match") or {}).get("pickup_name"), (m.get("match") or {}).get("delivery_name"),
                        (m.get("match") or {}).get("to_pickup_km"), (m.get("match") or {}).get("empty_km_reduced"),
                        (m.get("match") or {}).get("fill_after"), (m.get("match") or {}).get("revenue_add"),
                        (m.get("match") or {}).get("cost_add"), (m.get("match") or {}).get("profit_add"),
                        (m.get("match") or {}).get("score"), (m.get("match") or {}).get("decision")])
    bh = ["STT", "Xe", "Tuyến", "Đơn ghép", "Điểm lấy", "Điểm trả", "Đến điểm lấy (km)",
          "Km rỗng giảm", "Fill chiều về %", "Doanh thu thêm", "Chi phí thêm", "Lợi nhuận thêm",
          "Điểm phù hợp", "Quyết định"]
    ws_b, b_hr, b_d0, b_d1, b_cl = _sheet(
        wb, "Đơn bổ sung & Backhaul", "GHÉP CHUYẾN QUAY ĐẦU (BACKHAUL)",
        f"Đề xuất ghép chiều về · Cập nhật lần cuối: {stamp}",
        bh, bk_rows, widths=[5, 8, 14, 12, 20, 20, 13, 11, 12, 13, 13, 13, 10, 13],
        money_cols=(10, 11, 12), pct_cols=(9,), num_cols=(7, 8, 13), center_cols=(1,),
        table_name="tbl_backhaul")
    if bk_rows:
        ws_b.conditional_formatting.add(f"L{b_d0}:L{b_d1}",
            DataBarRule(start_type="min", end_type="max", color=GREEN))
        ws_b.conditional_formatting.add(f"H{b_d0}:H{b_d1}",
            DataBarRule(start_type="min", end_type="max", color=BLUE))

    # ---------- SHEET: TÀI CHÍNH P&L ----------
    pl_rows = []
    for i, p in enumerate(fin.get("per_route", []), 1):
        pl_rows.append([i, p.get("vehicle_id"), p.get("vehicle_type"), p.get("revenue_total"),
                        p.get("fuel"), p.get("toll"), p.get("driver_cost"), p.get("vehicle_cost"),
                        p.get("handling"), p.get("empty_cost"), p.get("overhead"),
                        p.get("total_cost"), p.get("profit"), p.get("margin")])
    ph = ["STT", "Mã xe", "Loại xe", "Doanh thu", "Nhiên liệu", "Cầu đường", "Tài xế",
          "Chi phí xe", "Bốc xếp", "Chạy rỗng", "Quản lý", "Tổng chi phí", "Lợi nhuận", "Biên %"]
    ws_p, p_hr, p_d0, p_d1, p_cl = _sheet(
        wb, "Tài chính P&L", "BÁO CÁO TÀI CHÍNH — LỢI NHUẬN KỲ VỌNG (P&L)",
        f"Biên kế hoạch: {kpi.get('margin','—')}% · Sau xử lý sự cố: {kpi.get('margin_after','—')}% · "
        f"Giá nhiên liệu: {int(_f(fuelc.get('gia_nhien_lieu'))):,}đ/lít ({fuelc.get('nguon','—')}) · "
        f"Cập nhật lần cuối: {stamp}".replace(",", "."),
        ph, pl_rows, widths=[5, 11, 10, 14, 12, 11, 12, 12, 11, 11, 11, 14, 14, 9],
        money_cols=(4, 5, 6, 7, 8, 9, 10, 11, 12, 13), pct_cols=(14,), center_cols=(1,),
        table_name="tbl_finance")
    if pl_rows:
        ws_p.conditional_formatting.add(f"N{p_d0}:N{p_d1}",
            ColorScaleRule(start_type="num", start_value=0, start_color="F4B9B5",
                           mid_type="num", mid_value=15, mid_color="FFE08A",
                           end_type="num", end_value=25, end_color="A8D5B5"))
        ws_p.conditional_formatting.add(f"M{p_d0}:M{p_d1}",
            DataBarRule(start_type="min", end_type="max", color=GREEN))
        ws_p.conditional_formatting.add(f"L{p_d0}:L{p_d1}",
            DataBarRule(start_type="min", end_type="max", color="C9A0A0"))
        # dòng TỔNG
        tot = fin.get("totals", {})
        tr = p_d1 + 1
        labels = ["TỔNG", "", "", tot.get("revenue_total"), tot.get("fuel"), tot.get("toll"),
                  tot.get("driver_cost"), tot.get("vehicle_cost"), tot.get("handling"),
                  tot.get("empty_cost"), tot.get("overhead"), tot.get("total_cost"),
                  tot.get("profit"), tot.get("margin")]
        for j, val in enumerate(labels, 1):
            cell = ws_p.cell(row=tr, column=j, value=val)
            cell.font = Font(name=FONT, size=11, bold=True, color=NAVY)
            cell.fill = _fill("FCE4E4")
            cell.border = BORDER
            if j in (4, 5, 6, 7, 8, 9, 10, 11, 12, 13):
                cell.number_format = VND
                cell.alignment = Alignment(horizontal="right")
            elif j == 14:
                cell.number_format = PCT
                cell.alignment = Alignment(horizontal="right")

    # ---------- SHEET: SỰ CỐ ----------
    inc_rows = []
    for i, c in enumerate(payload.get("incidents", []), 1):
        inc_rows.append([i, c.get("case_id"), c.get("order_id"), c.get("event_type"),
                         c.get("priority", "—"), c.get("ts", ""), c.get("vehicle", "—"),
                         c.get("route_id", "—"), c.get("status"), c.get("decision", "—")])
    ih = ["STT", "Mã sự cố", "Mã đơn", "Loại sự cố", "Ưu tiên", "Thời điểm", "Xe",
          "Tuyến", "Trạng thái", "Phương án chọn"]
    _sheet(wb, "Sự cố trong ngày", "SỰ CỐ TRONG NGÀY", f"Cập nhật lần cuối: {stamp}",
           ih, inc_rows, widths=[5, 14, 12, 20, 10, 18, 12, 10, 16, 30], center_cols=(1, 5),
           table_name="tbl_incidents", tab_color=RED_D)

    # ---------- SHEET: NHẬT KÝ ----------
    log_rows = []
    for i, e in enumerate(payload.get("log", []), 1):
        log_rows.append([i, e.get("ts"), e.get("case_id"), e.get("source"), e.get("order_id"),
                         e.get("event_type"), e.get("decision"), e.get("vehicle"),
                         e.get("status"), e.get("note", "")])
    lh = ["STT", "Thời điểm", "Mã sự cố", "Nguồn", "Đơn", "Loại", "Phương án", "Xe",
          "Trạng thái", "Ghi chú"]
    _sheet(wb, "Nhật ký điều phối", "NHẬT KÝ & LƯU VẾT ĐIỀU PHỐI",
           f"Phục vụ kiểm toán nội bộ / Close & Learn · Cập nhật lần cuối: {stamp}",
           lh, log_rows, widths=[5, 18, 14, 18, 12, 18, 30, 11, 16, 26], center_cols=(1,),
           table_name="tbl_log", tab_color=SLATE)

    # ---------- SHEET HƯỚNG DẪN ----------
    _build_guide(wb)

    # Xóa sheet rỗng mặc định để file chỉ còn các sheet dữ liệu + hướng dẫn.
    try:
        if _default_ws in wb.worksheets and len(wb.worksheets) > 1:
            wb.remove(_default_ws)
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_guide(wb):
    ws = wb.create_sheet("Hướng dẫn")
    ws.sheet_properties.tabColor = "A8B4CC"
    ws.sheet_view.showGridLines = False
    lines = [
        ("HƯỚNG DẪN SỬ DỤNG DỮ LIỆU BÁO CÁO", True, NAVY, 14),
        ("", False, INK, 11),
        ("• File chỉ gồm các sheet DỮ LIỆU (Kế hoạch tuyến, Đơn hàng, Xe & tài xế, Backhaul,", False, INK, 11),
        ("  Tài chính P&L, Sự cố, Nhật ký) — đã bỏ sheet Dashboard để tránh lỗi định dạng.", False, INK, 11),
        ("• Mỗi sheet dữ liệu là một Excel Table (có nút lọc, banded rows).", False, INK, 11),
        ("", False, INK, 11),
        ("TẠO PIVOTTABLE / SLICER / TIMELINE (native):", True, NAVY, 12),
        ("1) Mở sheet dữ liệu (vd 'Tài chính P&L' hoặc 'Kế hoạch tuyến').", False, INK, 11),
        ("2) Bấm vào bảng → ribbon Insert ▸ PivotTable (bảng đã đặt tên: tbl_finance, tbl_routes, tbl_orders, tbl_backhaul...).", False, INK, 11),
        ("3) Trong PivotTable: PivotTable Analyze ▸ Insert Slicer (lọc theo Tuyến/Loại xe...) và Insert Timeline (theo thời gian).", False, INK, 11),
        ("4) Tạo PivotChart: PivotTable Analyze ▸ PivotChart.", False, INK, 11),
        ("", False, INK, 11),
        ("GHI CHÚ: Slicer/Timeline/PivotTable là tính năng native của Excel, được bật sẵn nhờ dữ liệu đã chuẩn hóa thành Table.", False, SLATE, 10),
        ("Giá nhiên liệu là giá hệ thống tự đề xuất; doanh nghiệp có thể chỉnh tay trong sheet Tài chính P&L.", False, SLATE, 10),
    ]
    for i, (txt, big, color, size) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=txt)
        c.font = Font(name=FONT, size=size, bold=big, color=color)
    ws.column_dimensions["A"].width = 120
